"""
Kinder Morgan Pipeline Data Fetcher — 26 Pipelines
Automates IOC, Unsub, and Locations downloads from pipeline2.kindermorgan.com
via ASP.NET ViewState postback. No WAF, no browser session required.

Runs weekly via GitHub Actions (km-refresh.yml).
"""

import requests
import re
import csv
import json
import os
import io
import sys
import time
import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data')
OUTPUT_FILE = os.path.join(DATA_DIR, 'gas_interconnects.json')
TRACKER_FILE = os.path.join(DATA_DIR, 'corridor_pipeline_tracker.json')
COUNTY_CACHE = os.path.join(DATA_DIR, 'gas_county_coords.json')
TODAY = datetime.now().strftime('%Y-%m-%d')

UA = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'

# ============================================================
# 26 KM PIPELINE DEFINITIONS
# ============================================================

KM_PIPELINES = [
    # Interstate (18)
    {'code': 'TGP',  'name': 'Tennessee Gas Pipeline Company', 'short': 'Tennessee Gas', 'type': 'interstate', 'states': ['TX','LA','MS','AL','TN','KY','OH','PA','NJ','NY','CT','MA','NH']},
    {'code': 'NGPL', 'name': 'Natural Gas Pipeline Company of America', 'short': 'NGPL', 'type': 'interstate', 'states': ['TX','LA','AR','OK','KS','NE','MO','IA','IL','IN','WI']},
    {'code': 'EPNG', 'name': 'El Paso Natural Gas Company', 'short': 'El Paso', 'type': 'interstate', 'states': ['TX','NM','AZ','CA','NV','CO','UT']},
    {'code': 'SNG',  'name': 'Southern Natural Gas Company, LLC', 'short': 'Southern Natural', 'type': 'interstate', 'states': ['TX','LA','MS','AL','GA','SC']},
    {'code': 'CIG',  'name': 'Colorado Interstate Gas Company, LLC', 'short': 'Colorado Interstate', 'type': 'interstate', 'states': ['WY','CO']},
    {'code': 'MEP',  'name': 'Midcontinent Express Pipeline', 'short': 'Midcontinent Express', 'type': 'interstate', 'states': ['OK','TX','LA','MS','AL']},
    {'code': 'KMLP', 'name': 'Kinder Morgan Louisiana Pipeline', 'short': 'KM Louisiana', 'type': 'interstate', 'states': ['LA']},
    {'code': 'KMIL', 'name': 'KM Illinois Pipeline', 'short': 'KM Illinois', 'type': 'interstate', 'states': ['IL']},
    {'code': 'WIC',  'name': 'Wyoming Interstate Pipeline', 'short': 'WIC', 'type': 'interstate', 'states': ['WY','CO','UT']},
    {'code': 'TCP',  'name': 'TransColorado Gas Transmission', 'short': 'TransColorado', 'type': 'interstate', 'states': ['CO']},
    {'code': 'CP',   'name': 'Cheyenne Plains Gas Pipeline', 'short': 'Cheyenne Plains', 'type': 'interstate', 'states': ['CO','KS']},
    {'code': 'MOPC', 'name': 'Mojave Pipeline', 'short': 'Mojave', 'type': 'interstate', 'states': ['AZ','CA']},
    {'code': 'SGP',  'name': 'Sierrita Gas Pipeline', 'short': 'Sierrita', 'type': 'interstate', 'states': ['AZ']},
    {'code': 'EEC',  'name': 'Elba Express Company', 'short': 'Elba Express', 'type': 'interstate', 'states': ['GA','SC']},
    {'code': 'STAG', 'name': 'Stagecoach Pipeline & Storage', 'short': 'Stagecoach', 'type': 'interstate', 'states': ['NY','PA']},
    {'code': 'ARLS', 'name': 'Arlington Storage Company', 'short': 'Arlington Storage', 'type': 'interstate', 'states': ['NY']},
    {'code': 'YGS',  'name': 'Young Gas Storage', 'short': 'Young Gas Storage', 'type': 'interstate', 'states': ['CO']},
    {'code': 'TTP',  'name': 'Twin Tier Pipeline', 'short': 'Twin Tier', 'type': 'interstate', 'states': ['NY','PA']},
    # Intrastate (8)
    {'code': 'KMTP', 'name': 'Kinder Morgan Texas Pipeline', 'short': 'KM Texas', 'type': 'intrastate', 'states': ['TX']},
    {'code': 'KMTJ', 'name': 'Kinder Morgan Tejas Pipeline', 'short': 'KM Tejas', 'type': 'intrastate', 'states': ['TX']},
    {'code': 'KMNT', 'name': 'Kinder Morgan North Texas Pipeline', 'short': 'KM North Texas', 'type': 'intrastate', 'states': ['TX']},
    {'code': 'GCX',  'name': 'Gulf Coast Express Pipeline', 'short': 'GCX', 'type': 'intrastate', 'states': ['TX']},
    {'code': 'PHP',  'name': 'Permian Highway Pipeline', 'short': 'PHP', 'type': 'intrastate', 'states': ['TX']},
    {'code': 'KMBP', 'name': 'Kinder Morgan Border Pipeline', 'short': 'KM Border', 'type': 'intrastate', 'states': ['TX']},
    {'code': 'NETM', 'name': 'NET Mexico Pipeline', 'short': 'NET Mexico', 'type': 'intrastate', 'states': ['TX']},
    {'code': 'KMEF', 'name': 'Eagle Ford Midstream', 'short': 'Eagle Ford', 'type': 'intrastate', 'states': ['TX']},
]

# Map KM short names to tracker pipeline_name values (for existing entries)
KM_TRACKER_MAP = {
    'Tennessee Gas': 'Tennessee Gas Pipeline (TGP)',
    'NGPL': 'NGPL',
    'El Paso': 'El Paso Natural Gas',
    'Southern Natural': 'Southern Natural Gas (SNG)',
    'Colorado Interstate': 'Colorado Interstate Gas (CIG)',
    'Midcontinent Express': 'Midcontinent Express (MEP)',
    'KM Louisiana': 'Kinder Morgan Louisiana Pipeline',
    'KM Illinois': 'KM Illinois Pipeline',
    'WIC': 'Wyoming Interstate',
    'Mojave': 'Mojave Pipeline',
    'TransColorado': 'TransColorado',
}


# ============================================================
# ASP.NET POSTBACK HELPERS
# ============================================================

def extract_hidden_fields(html):
    """Extract ASP.NET hidden form fields (__VIEWSTATE, etc.)."""
    fields = {}
    for m in re.finditer(r'<input[^>]*type="hidden"[^>]*name="([^"]+)"[^>]*value="([^"]*)"', html):
        fields[m.group(1)] = m.group(2)
    # Also try reversed attribute order
    for m in re.finditer(r'<input[^>]*value="([^"]*)"[^>]*type="hidden"[^>]*name="([^"]+)"', html):
        fields[m.group(2)] = m.group(1)
    return fields


def km_session():
    s = requests.Session()
    s.headers['User-Agent'] = UA
    return s


def is_xlsx(resp):
    ct = resp.headers.get('Content-Type', '').lower()
    return 'excel' in ct or 'octet' in ct or resp.content[:2] == b'PK'


def is_csv_resp(resp):
    ct = resp.headers.get('Content-Type', '').lower()
    return 'text' in ct or 'csv' in ct


# ============================================================
# DATA FETCHERS
# ============================================================

def fetch_ioc(code):
    """Download IOC xlsx via ASP.NET postback. Returns list of row dicts."""
    s = km_session()
    url = f'https://pipeline2.kindermorgan.com/IndexOfCust/IOC.aspx?code={code}'
    r = s.get(url, timeout=30)

    hidden = extract_hidden_fields(r.text)
    form_data = dict(hidden)
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$DownloadDDL'] = 'EXCEL'
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload.x'] = '15'
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload.y'] = '15'

    r2 = s.post(url, data=form_data, timeout=60)

    if is_xlsx(r2):
        return parse_ioc_xlsx(r2.content)
    print(f"    [IOC] Not xlsx — Content-Type: {r2.headers.get('Content-Type','?')}")
    return {'contracts': [], 'by_point': {}}


def fetch_unsub(code):
    """Download Unsubscribed Capacity xlsx via ASP.NET postback."""
    s = km_session()
    url = f'https://pipeline2.kindermorgan.com/Capacity/UnsubscribedPoint.aspx?code={code}'
    r = s.get(url, timeout=30)

    hidden = extract_hidden_fields(r.text)
    form_data = dict(hidden)
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$DownloadDDL'] = 'EXCEL'
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload.x'] = '15'
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload.y'] = '15'

    r2 = s.post(url, data=form_data, timeout=60)

    if is_xlsx(r2):
        return parse_unsub_xlsx(r2.content)
    print(f"    [Unsub] Not xlsx — Content-Type: {r2.headers.get('Content-Type','?')}")
    return []


def fetch_locations(code):
    """Download Locations CSV via ASP.NET postback."""
    s = km_session()
    url = f'https://pipeline2.kindermorgan.com/LocationDataDownload/LocDataDwnld.aspx?code={code}'
    r = s.get(url, timeout=30)

    hidden = extract_hidden_fields(r.text)
    form_data = dict(hidden)
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$DownloadDDL'] = 'CSV'
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload.x'] = '15'
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload.y'] = '15'
    form_data['ctl00$hdnIsDownload'] = 'true'

    r2 = s.post(url, data=form_data, timeout=60)

    if is_csv_resp(r2):
        rows = list(csv.DictReader(io.StringIO(r2.text)))
        # Index by Loc ID
        locs = {}
        for row in rows:
            loc_id = str(row.get('Loc', '')).strip()
            if loc_id:
                locs[loc_id] = row
        return locs
    print(f"    [Locs] Not CSV — Content-Type: {r2.headers.get('Content-Type','?')}")
    return {}


# ============================================================
# PARSERS
# ============================================================

def parse_int_safe(val):
    """Parse integer from possibly comma-formatted string."""
    if val is None:
        return 0
    try:
        return int(str(val).replace(',', '').strip())
    except (ValueError, TypeError):
        return 0


def parse_ioc_xlsx(content):
    """Parse IOC Excel into contracts list and per-point aggregates."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (contains "Shipper Name")
    header_idx = None
    for i, row in enumerate(rows):
        if row and any('Shipper' in str(c) for c in row if c):
            header_idx = i
            break
    if header_idx is None:
        return {'contracts': [], 'by_point': {}}

    headers = [str(h).strip() if h else f'col{j}' for j, h in enumerate(rows[header_idx])]
    data = [dict(zip(headers, row)) for row in rows[header_idx + 1:] if row and row[0] is not None]

    # Aggregate by contract and by point
    contracts_map = {}  # contract_no -> contract dict
    by_point = defaultdict(lambda: {'firm_mdq': 0, 'expiring_2yr': 0, 'num_contracts': 0, 'shippers': set()})
    cutoff = datetime.now() + timedelta(days=730)

    all_shippers = set()

    for d in data:
        shipper = str(d.get('Shipper Name', '')).strip()
        contract_no = str(d.get('Contract No.', d.get('Contract No', ''))).strip()
        rate = str(d.get('Rate Sched', '')).strip()
        mdq = parse_int_safe(d.get('MDQ') or d.get('PT MDQ') or 0)
        pt_mdq = parse_int_safe(d.get('PT MDQ') or d.get('MDQ') or 0)
        point_id = str(d.get('Point ID', d.get('Pt ID CD', ''))).strip()
        point_name = str(d.get('Pt Name', '')).strip()
        exp_date = d.get('Contract Expiration Date', '')
        eff_date = d.get('Contract Effective Date', '')
        zone = str(d.get('Zone Name', '')).strip()

        if not point_id or pt_mdq == 0:
            continue

        all_shippers.add(shipper)

        # Aggregate per contract
        if contract_no and contract_no not in contracts_map:
            contracts_map[contract_no] = {
                'shipper': shipper,
                'rate_schedule': rate,
                'contract_id': contract_no,
                'begin_date': str(eff_date).strip()[:10] if eff_date else '',
                'end_date': str(exp_date).strip()[:10] if exp_date else '',
                'mdq_dth': parse_int_safe(d.get('MDQ', 0)),
                'points': [],
            }
        if contract_no in contracts_map:
            contracts_map[contract_no]['points'].append({
                'loc_id': point_id,
                'loc_nm': point_name,
                'qty': pt_mdq,
            })

        # Aggregate per point
        by_point[point_id]['num_contracts'] += 1
        by_point[point_id]['shippers'].add(shipper)
        if 'FT' in rate.upper():
            by_point[point_id]['firm_mdq'] += pt_mdq

        if exp_date:
            try:
                if isinstance(exp_date, datetime):
                    ed = exp_date
                else:
                    ed = datetime.strptime(str(exp_date).strip()[:10], '%m/%d/%Y')
                if ed <= cutoff:
                    by_point[point_id]['expiring_2yr'] += pt_mdq
            except (ValueError, TypeError):
                pass

    # Convert sets to counts
    by_point_out = {}
    for loc_id, info in by_point.items():
        by_point_out[loc_id] = {
            'firm_mdq': info['firm_mdq'],
            'expiring_2yr': info['expiring_2yr'],
            'num_contracts': info['num_contracts'],
            'num_shippers': len(info['shippers']),
        }

    contracts = list(contracts_map.values())
    total_mdq = sum(c['mdq_dth'] for c in contracts)

    return {
        'contracts': contracts,
        'by_point': by_point_out,
        'total_mdq': total_mdq,
        'num_contracts': len(contracts),
        'num_shippers': len(all_shippers),
    }


def parse_unsub_xlsx(content):
    """Parse Unsubscribed Capacity Excel."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (contains "Loc" and "Unsubscribed")
    header_idx = None
    for i, row in enumerate(rows):
        if row and any('Loc' == str(c).strip() for c in row if c):
            header_idx = i
            break
    if header_idx is None:
        return []

    headers = [str(h).strip() if h else f'col{j}' for j, h in enumerate(rows[header_idx])]
    result = []
    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        loc_id = str(row[0]).strip()
        # Skip footer rows
        if 'Row Count' in loc_id or 'Comments' in loc_id or not loc_id:
            continue
        d = dict(zip(headers, row))
        unsub_val = parse_int_safe(d.get('Unsubscribed Capacity', 0))
        if unsub_val > 0:
            result.append({
                'Loc': loc_id,
                'Loc_Name': str(d.get('Loc Name', '')).strip(),
                'Loc_Purp_Desc': str(d.get('Loc Zn', '')).strip(),
                'Unsubscribed_Capacity': unsub_val,
            })
    return result


# ============================================================
# COUNTY CENTROID GEOCODING
# ============================================================

def load_county_coords():
    """Load cached county centroid coordinates."""
    if os.path.exists(COUNTY_CACHE):
        with open(COUNTY_CACHE) as f:
            return json.load(f)
    return {}


def save_county_coords(coords):
    with open(COUNTY_CACHE, 'w') as f:
        json.dump(coords, f)


def geocode_county(county, state):
    """Geocode a county via Census.gov geocoder (free, no key needed)."""
    try:
        url = 'https://geocoding.geo.census.gov/geocoder/geographies/onelineaddress'
        params = {
            'address': f'{county} County, {state}',
            'benchmark': 'Public_AR_Current',
            'vintage': 'Current_Current',
            'format': 'json',
        }
        r = requests.get(url, params=params, timeout=10)
        d = r.json()
        matches = d.get('result', {}).get('addressMatches', [])
        if matches:
            coords = matches[0].get('coordinates', {})
            return coords.get('y'), coords.get('x')
    except Exception:
        pass

    # Fallback: use a static lookup for common states
    # (Census geocoder sometimes fails for county-level queries)
    return None, None


# ============================================================
# MAIN PIPELINE PROCESSING
# ============================================================

def process_pipeline(pl, county_coords):
    """Fetch all data for one KM pipeline and return gas_interconnects entry."""
    code = pl['code']
    short = pl['short']

    print(f"\n  [{code}] {short}...")

    # Fetch all three data types
    try:
        ioc_data = fetch_ioc(code)
    except Exception as e:
        print(f"    IOC error: {e}")
        ioc_data = {'contracts': [], 'by_point': {}}

    time.sleep(1)

    try:
        unsub_data = fetch_unsub(code)
    except Exception as e:
        print(f"    Unsub error: {e}")
        unsub_data = []

    time.sleep(1)

    try:
        loc_data = fetch_locations(code)
    except Exception as e:
        print(f"    Locations error: {e}")
        loc_data = {}

    time.sleep(1)

    # Build points from locations data (primary source for point list)
    # Fall back to IOC point IDs if no locations available
    points = []
    point_ids_seen = set()

    by_point = ioc_data.get('by_point', {})

    for loc_id, loc in loc_data.items():
        county = str(loc.get('Loc Cnty', '')).strip()
        state = str(loc.get('Loc St Abbrev', '')).strip()
        flow = str(loc.get('Dir Flo', '')).strip()
        loc_type = str(loc.get('Loc Type Ind', '')).strip()
        connected = str(loc.get('Up/Dn Name', '')).strip()[:50]
        loc_name = str(loc.get('Loc Name', '')).strip()

        ptype = 'delivery'
        if flow == 'R':
            ptype = 'receipt'
        elif flow == 'B':
            ptype = 'bidirectional'

        # Geocode
        lat, lng = None, None
        if county and state:
            key = f"{county.upper()}|{state}"
            if key in county_coords:
                lat, lng = county_coords[key].get('lat'), county_coords[key].get('lng')

        ioc_info = by_point.get(loc_id, {})

        pt = {
            'id': loc_id,
            'name': loc_name[:50],
            'type': ptype,
            'county': county,
            'state': state,
            'design': 0,
            'scheduled': 0,
            'available': 0,
            'utilization': 0,
            'connected': connected,
            'firm_contracted': ioc_info.get('firm_mdq', 0),
            'num_contracts': ioc_info.get('num_contracts', 0),
            'num_shippers': ioc_info.get('num_shippers', 0),
            'expiring_2yr': ioc_info.get('expiring_2yr', 0),
        }
        if lat and lng:
            pt['lat'] = lat
            pt['lng'] = lng
            pt['loc_accuracy'] = 'county_centroid'

        points.append(pt)
        point_ids_seen.add(loc_id)

    # Add any IOC-only points not in locations
    for loc_id, info in by_point.items():
        if loc_id not in point_ids_seen:
            points.append({
                'id': loc_id,
                'name': '',
                'type': 'other',
                'county': '',
                'state': '',
                'design': 0,
                'scheduled': 0,
                'available': 0,
                'utilization': 0,
                'connected': '',
                'firm_contracted': info.get('firm_mdq', 0),
                'num_contracts': info.get('num_contracts', 0),
                'num_shippers': info.get('num_shippers', 0),
                'expiring_2yr': info.get('expiring_2yr', 0),
            })

    # Collect new counties that need geocoding
    new_counties = []
    for pt in points:
        county = pt.get('county', '').upper()
        state = pt.get('state', '')
        if county and state:
            key = f"{county}|{state}"
            if key not in county_coords:
                new_counties.append(key)

    # Build IOC totals
    contracts = ioc_data.get('contracts', [])
    ioc_totals = {
        'firm_mdq': ioc_data.get('total_mdq', 0),
        'num_contracts': ioc_data.get('num_contracts', 0),
        'num_shippers': ioc_data.get('num_shippers', 0),
    }

    entry = {
        'name': pl['name'],
        'short': short,
        'updated': TODAY,
        'points': points,
        'unsub_points': unsub_data,
        'ioc_totals': ioc_totals,
    }

    print(f"    {len(points)} pts, {len(unsub_data)} unsub, {ioc_totals['num_contracts']} IOC, {ioc_totals['num_shippers']} shippers, {ioc_totals['firm_mdq']:,} MDQ")

    return entry, list(set(new_counties))


# ============================================================
# GAS_INTERCONNECTS MERGE
# ============================================================

def merge_into_gas_interconnects(new_entries):
    """Merge KM pipeline entries into gas_interconnects.json, replacing existing by short name."""
    if os.path.exists(OUTPUT_FILE):
        with open(OUTPUT_FILE) as f:
            gi = json.load(f)
    else:
        gi = {'pipelines': []}

    # Build set of short names we're updating
    km_shorts = {e['short'] for e in new_entries}

    # Remove existing entries for these pipelines
    gi['pipelines'] = [p for p in gi['pipelines'] if p.get('short') not in km_shorts]

    # Add new entries
    gi['pipelines'].extend(new_entries)

    with open(OUTPUT_FILE, 'w') as f:
        json.dump(gi, f)

    print(f"\ngas_interconnects.json: {len(gi['pipelines'])} total pipelines")


# ============================================================
# TRACKER UPDATE
# ============================================================

def update_tracker(results):
    """Update corridor_pipeline_tracker.json for all KM pipelines."""
    if not os.path.exists(TRACKER_FILE):
        print("Tracker file not found — skipping")
        return

    with open(TRACKER_FILE) as f:
        tracker = json.load(f)

    existing_names = {e.get('pipeline_name', ''): i for i, e in enumerate(tracker['gas_pipelines'])}

    updated = 0
    added = 0

    for pl in KM_PIPELINES:
        short = pl['short']
        tracker_name = KM_TRACKER_MAP.get(short)

        # Find existing entry
        entry_idx = None
        if tracker_name and tracker_name in existing_names:
            entry_idx = existing_names[tracker_name]
        else:
            # Try matching by pipeline name directly
            for name_key in [pl['name'], short]:
                if name_key in existing_names:
                    entry_idx = existing_names[name_key]
                    break

        if entry_idx is not None:
            # Update existing entry
            entry = tracker['gas_pipelines'][entry_idx]

            # Update IOC
            if 'ioc' in entry:
                entry['ioc']['access_method'] = 'weekly_auto'
                entry['ioc']['platform'] = 'kinder_morgan_postback'
                entry['ioc']['last_refreshed'] = TODAY
                entry['ioc']['refresh_frequency_days'] = 7
            else:
                entry['ioc'] = {
                    'status': 'captured',
                    'access_method': 'weekly_auto',
                    'platform': 'kinder_morgan_postback',
                    'url': f'https://pipeline2.kindermorgan.com/IndexOfCust/IOC.aspx?code={pl["code"]}',
                    'last_refreshed': TODAY,
                    'refresh_frequency_days': 7,
                }

            # Update Unsub
            if 'unsub' in entry:
                entry['unsub']['access_method'] = 'weekly_auto'
                entry['unsub']['platform'] = 'kinder_morgan_postback'
                entry['unsub']['last_refreshed'] = TODAY
            else:
                entry['unsub'] = {
                    'status': 'captured',
                    'access_method': 'weekly_auto',
                    'platform': 'kinder_morgan_postback',
                    'last_refreshed': TODAY,
                }

            # Update locations
            if 'locations' not in entry:
                entry['locations'] = {}
            entry['locations']['status'] = 'captured'
            entry['locations']['access_method'] = 'weekly_auto'

            updated += 1
        else:
            # Add new tracker entry
            new_entry = {
                'pipeline_name': pl['name'],
                'pipeline_id': pl['code'].lower(),
                'operator': 'Kinder Morgan',
                'parent': 'Kinder Morgan',
                'regulation': pl['type'],
                'hifld_points': 0,
                'type': 'pipeline',
                'ioc': {
                    'status': 'captured',
                    'access_method': 'weekly_auto',
                    'platform': 'kinder_morgan_postback',
                    'url': f'https://pipeline2.kindermorgan.com/IndexOfCust/IOC.aspx?code={pl["code"]}',
                    'last_refreshed': TODAY,
                    'refresh_frequency_days': 7,
                },
                'unsub': {
                    'status': 'captured',
                    'access_method': 'weekly_auto',
                    'platform': 'kinder_morgan_postback',
                    'last_refreshed': TODAY,
                },
                'locations': {
                    'status': 'captured',
                    'access_method': 'weekly_auto',
                    'geocode_method': 'county_centroid',
                },
                'notes': f'KM portal code={pl["code"]}. States: {", ".join(pl["states"])}.',
            }
            tracker['gas_pipelines'].append(new_entry)
            added += 1

    tracker['_summary']['total_pipelines'] = len(tracker['gas_pipelines'])

    with open(TRACKER_FILE, 'w') as f:
        json.dump(tracker, f, indent=2)

    print(f"\nTracker: {updated} updated, {added} added, {len(tracker['gas_pipelines'])} total")


# ============================================================
# MAIN
# ============================================================

if __name__ == '__main__':
    print(f"=== KM Pipeline Refresh: {TODAY} ===")
    print(f"Processing {len(KM_PIPELINES)} pipelines\n")

    county_coords = load_county_coords()
    all_entries = []
    all_new_counties = []
    success = 0
    failed = 0

    for pl in KM_PIPELINES:
        try:
            entry, new_counties = process_pipeline(pl, county_coords)
            all_entries.append(entry)
            all_new_counties.extend(new_counties)
            success += 1
        except Exception as e:
            print(f"  FAILED {pl['short']}: {e}")
            failed += 1
        time.sleep(2)

    # Geocode any new counties
    all_new_counties = list(set(all_new_counties))
    if all_new_counties:
        print(f"\nGeocoding {len(all_new_counties)} new counties...")
        geocoded = 0
        for key in all_new_counties:
            county, state = key.split('|')
            lat, lng = geocode_county(county, state)
            if lat and lng:
                county_coords[key] = {'lat': lat, 'lng': lng}
                geocoded += 1
            time.sleep(0.5)
        save_county_coords(county_coords)
        print(f"  Geocoded {geocoded}/{len(all_new_counties)}")

        # Re-apply coordinates to points
        for entry in all_entries:
            for pt in entry['points']:
                county = pt.get('county', '').upper()
                state = pt.get('state', '')
                if county and state and 'lat' not in pt:
                    key = f"{county}|{state}"
                    if key in county_coords:
                        pt['lat'] = county_coords[key]['lat']
                        pt['lng'] = county_coords[key]['lng']
                        pt['loc_accuracy'] = 'county_centroid'

    # Merge into gas_interconnects.json
    merge_into_gas_interconnects(all_entries)

    # Update tracker
    update_tracker(all_entries)

    # Summary
    total_pts = sum(len(e['points']) for e in all_entries)
    total_unsub = sum(len(e.get('unsub_points', [])) for e in all_entries)
    total_ioc = sum(e.get('ioc_totals', {}).get('num_contracts', 0) for e in all_entries)
    print(f"\n=== Done: {success} OK, {failed} failed ===")
    print(f"Total: {total_pts} points, {total_unsub} unsub, {total_ioc} IOC contracts")
