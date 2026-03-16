"""
Gas Interconnect Data Fetcher
Pulls EBB capacity and IOC contract data from pipeline operator portals.
Runs daily via GitHub Actions to build 30-day rolling averages.

Platforms:
  - Kinder Morgan pipeline2 (El Paso, Tennessee Gas, NGPL, MEP, Southern Natural)
  - Williams 1line (Transco)
  - Enbridge rtba (Texas Eastern, Algonquin, Maritimes NE, East Tennessee)
  - Enbridge IOC CSV (Texas Eastern, Algonquin, East Tennessee, Maritimes NE)
  - Energy Transfer CSV (CenterPoint/EGT) — OAC via direct CSV download
  - TC Plus tcplus.com (Great Lakes, GTN, Tuscarora) — IOC only
"""

import requests
import re
import csv
import json
import os
import io
import sys
import time
import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict

GEOCODIO_KEY = os.environ.get('GEOCODIO_KEY', '')
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data')
HISTORY_FILE = os.path.join(DATA_DIR, 'gas_history.json')
COUNTY_CACHE = os.path.join(DATA_DIR, 'gas_county_coords.json')
HIFLD_CACHE = os.path.join(DATA_DIR, 'gas_hifld_points.json')
OUTPUT_FILE = os.path.join(DATA_DIR, 'gas_interconnects.json')

TODAY = datetime.now().strftime('%Y-%m-%d')
TODAY_MMDDYYYY = datetime.now().strftime('%m/%d/%Y')

# ============================================================
# KINDER MORGAN (ASP.NET ViewState POST)
# ============================================================

KM_PIPELINES = [
    {'code': 'EPNG', 'name': 'El Paso Natural Gas Company', 'short': 'El Paso'},
    {'code': 'TGP', 'name': 'Tennessee Gas Pipeline Company', 'short': 'Tennessee Gas'},
    {'code': 'NGPL', 'name': 'Natural Gas Pipeline Company of America', 'short': 'NGPL'},
    {'code': 'MEP', 'name': 'Midcontinent Express Pipeline', 'short': 'Midcontinent Express'},
    {'code': 'SNG', 'name': 'Southern Natural Gas Company, LLC', 'short': 'Southern Natural'},
    {'code': 'FGT', 'name': 'Florida Gas Transmission Company, LLC', 'short': 'Florida Gas'},
    {'code': 'CIG', 'name': 'Colorado Interstate Gas Company, LLC', 'short': 'Colorado Interstate'},
]

def fetch_km_capacity(code):
    """Fetch Operationally Available Capacity from KM portal."""
    s = requests.Session()
    s.headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    
    url = f'https://pipeline2.kindermorgan.com/Capacity/OpAvailPoint.aspx?code={code}'
    r = s.get(url)
    
    hidden = extract_hidden_fields(r.text)
    form_data = dict(hidden)
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$DownloadDDL'] = 'EXCEL'
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload.x'] = '10'
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload.y'] = '10'
    
    r2 = s.post(url, data=form_data)
    
    if 'excel' in r2.headers.get('Content-Type', '').lower() or r2.content[:2] == b'PK':
        return parse_km_xlsx(r2.content)
    return []


def fetch_km_locations(code):
    """Fetch Location Data Download from KM portal."""
    s = requests.Session()
    s.headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    
    url = f'https://pipeline2.kindermorgan.com/LocationDataDownload/LocDataDwnld.aspx?code={code}'
    r = s.get(url)
    
    hidden = extract_hidden_fields(r.text)
    form_data = dict(hidden)
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$DownloadDDL'] = 'CSV'
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload.x'] = '10'
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload.y'] = '10'
    
    r2 = s.post(url, data=form_data)
    
    if 'text' in r2.headers.get('Content-Type', ''):
        return {str(r['Loc']).strip(): r for r in csv.DictReader(io.StringIO(r2.text))}
    return {}


def fetch_km_ioc(code):
    """Fetch Index of Customers from KM portal."""
    s = requests.Session()
    s.headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    
    url = f'https://pipeline2.kindermorgan.com/IndexOfCust/IOC.aspx?code={code}'
    r = s.get(url)
    
    hidden = extract_hidden_fields(r.text)
    form_data = dict(hidden)
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$DownloadDDL'] = 'EXCEL'
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload.x'] = '10'
    form_data['ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload.y'] = '10'
    
    r2 = s.post(url, data=form_data)
    
    if 'excel' in r2.headers.get('Content-Type', '').lower() or 'octet' in r2.headers.get('Content-Type', ''):
        return parse_ioc_xlsx(r2.content)
    return {}


# ============================================================
# WILLIAMS (JSP Form POST + HTML Parse)
# ============================================================

WILLIAMS_PIPELINES = [
    {'buid': 80, 'name': 'Transcontinental Gas Pipe Line Company (Transco)', 'short': 'Transco'},
]

def fetch_williams_capacity(buid):
    """Fetch OAC from Williams 1line JSP portal."""
    s = requests.Session()
    s.headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    
    s.get(f'https://www.1line.williams.com/ebbCode/OACQueryRequest.jsp?BUID={buid}')
    s.post(f'https://www.1line.williams.com/ebbCode/OACQueryRequest.jsp?BUID={buid}', data={
        'tbGasFlowBeginDate': TODAY_MMDDYYYY,
        'tbGasFlowEndDate': TODAY_MMDDYYYY,
        'cycle': '2',
        'locationIDs': '',
        'reportType': 'OAC',
        'submitflag': 'true',
        'MapID': '0',
    })
    
    r = s.get('https://www.1line.williams.com/ebbCode/OACreport.jsp')
    return parse_williams_html(r.text)


# ============================================================
# ENBRIDGE (ASP.NET ViewState POST → CSV)
# ============================================================

ENBRIDGE_PIPELINES = [
    {'bu': 'TE', 'name': 'Texas Eastern Transmission, LP', 'short': 'Texas Eastern'},
    {'bu': 'AG', 'name': 'Algonquin Gas Transmission, LLC', 'short': 'Algonquin'},
    {'bu': 'MN', 'name': 'Maritimes & Northeast Pipeline, LLC', 'short': 'Maritimes NE'},
    {'bu': 'ET', 'name': 'East Tennessee Natural Gas, LLC', 'short': 'East Tennessee'},
]

# Enbridge bu codes map to different IOC CSV filename codes
ENBRIDGE_IOC_CODES = {
    'TE': 'TE',  # Texas Eastern
    'AG': 'AG',  # Algonquin
    'ET': 'ET',  # East Tennessee
    'MN': 'MN',  # Maritimes NE
}

def fetch_enbridge_capacity(bu_code):
    """Fetch OAC CSV from Enbridge rtba portal."""
    s = requests.Session()
    s.headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    
    r = s.get(f'https://rtba.enbridge.com/InformationalPosting/Default.aspx?Type=OA&bu={bu_code}')
    
    hidden = extract_hidden_fields(r.text)
    cycles = re.findall(r'<option[^>]*value="([^"]*)"', r.text)
    if not cycles:
        return []
    
    form_data = dict(hidden)
    form_data[f'ctl00$MainContent$ctl01$oaDefault$ddlSelector'] = cycles[0]
    form_data['__EVENTTARGET'] = 'ctl00$MainContent$ctl01$oaDefault$hlDown$LinkButton1'
    form_data['__EVENTARGUMENT'] = ''
    
    r2 = s.post(f'https://rtba.enbridge.com/InformationalPosting/Default.aspx?Type=OA&bu={bu_code}', data=form_data)
    
    if 'text/plain' in r2.headers.get('Content-Type', ''):
        return list(csv.DictReader(io.StringIO(r2.text)))
    return []


def fetch_enbridge_ioc(bu_code):
    """Fetch Index of Customers CSV from Enbridge Downloads endpoint.

    CSV format uses row types: H=header, D=contract detail, P=point detail.
    We parse D rows for shipper, rate schedule, dates, and MDQ.
    Returns pipeline-level IOC stats dict.
    """
    ioc_code = ENBRIDGE_IOC_CODES.get(bu_code, bu_code)
    url = f'https://infopost.enbridge.com/Downloads/IOC/{ioc_code}_IOC.csv'

    try:
        r = requests.get(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }, timeout=30)
        if r.status_code != 200:
            print(f"    IOC CSV returned {r.status_code} for {bu_code}")
            return {}
    except Exception as e:
        print(f"    IOC CSV fetch error for {bu_code}: {e}")
        return {}

    cutoff = datetime.now() + timedelta(days=730)
    firm_mdq = 0
    expiring_2yr = 0
    num_contracts = 0
    shippers = set()

    reader = csv.reader(io.StringIO(r.text))
    for row in reader:
        if not row or row[0].strip() != 'D':
            continue

        # D row fields (0-indexed): 0=type, 1=shipper, 2=DUNS, 3=affiliate,
        # 4=rate_schedule, 5=k_id, 6=k_beg_date, 7=k_end_date, 8=amendment,
        # 9=nego_rate, 10=MDQ, 11=storage_qty, 12=footnote
        if len(row) < 11:
            continue

        shipper = row[1].strip()
        rate_schedule = row[4].strip()
        k_end_str = row[7].strip()
        mdq_str = row[10].strip() if len(row) > 10 else '0'

        try:
            mdq = int(mdq_str.replace(',', ''))
        except (ValueError, AttributeError):
            mdq = 0

        if mdq == 0 or not shipper:
            continue

        num_contracts += 1
        shippers.add(shipper)

        if 'FT' in rate_schedule.upper():
            firm_mdq += mdq

        if k_end_str:
            try:
                ed = datetime.strptime(k_end_str.strip()[:10], '%m/%d/%Y')
                if ed <= cutoff:
                    expiring_2yr += mdq
            except (ValueError, IndexError):
                pass

    return {
        'firm_mdq': firm_mdq,
        'expiring_2yr': expiring_2yr,
        'num_contracts': num_contracts,
        'num_shippers': len(shippers),
    }


# ============================================================
# ENERGY TRANSFER (HTML Table Parse)
# ============================================================

ET_PIPELINES = [
    {'base': 'peplmessenger', 'asset': 'PEPL', 'name': 'Panhandle Eastern Pipe Line Company', 'short': 'Panhandle Eastern'},
    {'base': 'tgcmessenger', 'asset': 'TGC', 'name': 'Trunkline Gas Company, LLC', 'short': 'Trunkline'},
    {'base': 'rovermessenger', 'asset': 'ROVER', 'name': 'Rover Pipeline LLC', 'short': 'Rover'},
]

def fetch_et_capacity(base_domain, asset):
    """Fetch OAC from Energy Transfer messenger platform via HTML table parse."""
    s = requests.Session()
    s.headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    
    url = f'https://{base_domain}.energytransfer.com/ipost/capacity/operationally-available-by-location?asset={asset}'
    r = s.get(url, timeout=30)
    
    all_rows = re.findall(r'<tr[^>]*>(.*?)</tr>', r.text, re.DOTALL | re.I)
    data = []
    for row in all_rows:
        cells = re.findall(r'<td[^>]*>(.*?)</td>', row, re.DOTALL | re.I)
        clean = [re.sub(r'<[^>]+>', '', c).strip() for c in cells]
        if len(clean) >= 10:
            has_num = any(c.replace(',', '').replace('.', '').replace('-', '').isdigit() for c in clean[4:8] if c)
            if has_num:
                data.append({
                    'Loc': clean[0],
                    'Loc_Name': clean[1],
                    'Loc_Purp_Desc': clean[2],
                    'Total_Design_Capacity': clean[4],
                    'Operating_Capacity': clean[5],
                    'Total_Scheduled_Quantity': clean[6],
                    'Operationally_Available_Capacity': clean[7],
                    'Loc_Zn': clean[8] if len(clean) > 8 else '',
                    'Flow_Ind_Desc': clean[10] if len(clean) > 10 else '',
                    'State': clean[11] if len(clean) > 11 else '',
                    'County': clean[12] if len(clean) > 12 else '',
                    'Operator': clean[13] if len(clean) > 13 else '',
                })
    return data


def fetch_et_ioc(base_domain, asset):
    """Fetch Index of Customers from Energy Transfer messenger platform."""
    s = requests.Session()
    s.headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    
    url = f'https://{base_domain}.energytransfer.com/ipost/index-of-customers/index?asset={asset}'
    r = s.get(url, timeout=30)
    
    all_rows = re.findall(r'<tr[^>]*>(.*?)</tr>', r.text, re.DOTALL | re.I)
    
    cutoff = datetime.now() + timedelta(days=730)
    firm_mdq = 0
    expiring_2yr = 0
    num_contracts = 0
    shippers = set()
    
    for row in all_rows:
        cells = re.findall(r'<td[^>]*>(.*?)</td>', row, re.DOTALL | re.I)
        if len(cells) < 7:
            continue
        clean = [re.sub(r'<[^>]+>', '', c).strip() for c in cells]
        
        shipper = clean[0]
        rate = clean[3] if len(clean) > 3 else ''
        exp_date_str = clean[6] if len(clean) > 6 else ''
        mdq_str = clean[9] if len(clean) > 9 else '0'
        
        if not shipper or not clean[4]:  # need shipper and contract number
            continue
        
        try:
            mdq = int(mdq_str.replace(',', ''))
        except:
            mdq = 0
        
        if mdq == 0:
            continue
        
        num_contracts += 1
        shippers.add(shipper)
        
        if 'FT' in rate.upper() or 'EFT' in rate.upper() or 'NNS' in rate.upper():
            firm_mdq += mdq
        
        if exp_date_str:
            try:
                ed = datetime.strptime(exp_date_str.strip()[:10], '%m/%d/%Y')
                if ed <= cutoff:
                    expiring_2yr += mdq
            except:
                pass
    
    return {
        'firm_mdq': firm_mdq,
        'expiring_2yr': expiring_2yr,
        'num_contracts': num_contracts,
        'num_shippers': len(shippers),
    }


# ============================================================
# CENTERPOINT / EGT (Energy Transfer CSV downloads, no auth)
# Different CSV format from ET messenger HTML tables
# ============================================================

EGT_PIPELINES = [
    {'asset': 'EGT', 'name': 'CenterPoint Energy Gas Transmission Company', 'short': 'CenterPoint EGT'},
]


def fetch_egt_capacity(asset):
    """Fetch OAC + measuring point data from Energy Transfer CSV downloads.

    OAC CSV columns: FLOW MONTH, FLOW DAY, LOCATION, LOCATION NAME, POOLING AREA,
    LOC PURPOSE, LOC LABEL, ALL QTYS AVAILABLE, DESIGN CAPACITY, OPERATIONAL CAPACITY,
    SCHED QTY, AVAILABLE CAPACITY, FLOW DIR, IT INCLUDED
    Values are zero-padded integers (e.g. '000052000').

    Measuring points CSV provides STATE and COUNTY per location.
    """
    s = requests.Session()
    s.headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'

    # Fetch OAC capacity CSV (follows redirect)
    oac_url = f'https://pipelines.energytransfer.com/ipost/downloads/oper-avail-cap?asset={asset}'
    r = s.get(oac_url, timeout=30, allow_redirects=True)
    if r.status_code != 200:
        return [], {}

    oac_rows = list(csv.DictReader(io.StringIO(r.text)))

    # Fetch measuring points CSV for state/county
    mp_url = f'https://pipelines.energytransfer.com/ipost/downloads/measuring-point?asset={asset}'
    r2 = s.get(mp_url, timeout=30, allow_redirects=True)
    loc_map = {}
    if r2.status_code == 200 and len(r2.text) > 100:
        for row in csv.DictReader(io.StringIO(r2.text)):
            loc_id = row.get('LOCATION', '').strip()
            if loc_id:
                loc_map[loc_id] = {
                    'state': row.get('STATE', '').strip(),
                    'county': row.get('COUNTY', '').strip(),
                    'operator': row.get('LOCATION OPERATOR', '').strip(),
                }

    return oac_rows, loc_map


# ============================================================
# TC ENERGY (eConnects ReportViewer CSV)
# ============================================================

TC_PIPELINES = [
    {'asset': 3005, 'report': 'OperationallyAvailableCapacityANR', 'name': 'ANR Pipeline Company', 'short': 'ANR Pipeline'},
    {'asset': 51, 'report': 'OperationallyAvailableCapacity', 'name': 'Columbia Gas Transmission, LLC', 'short': 'Columbia Gas'},
    {'asset': 14, 'report': 'OperationallyAvailableCapacity', 'name': 'Columbia Gulf Transmission, LLC', 'short': 'Columbia Gulf'},
    {'asset': 3029, 'report': 'OperationallyAvailableCapacity', 'name': 'Northern Border Pipeline Company', 'short': 'Northern Border'},
]

def fetch_tc_capacity(asset_id, report_name):
    """Fetch OAC CSV from TC Energy eConnects ReportViewer."""
    s = requests.Session()
    s.headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    
    oac_url = f'https://ebb.tceconnects.com/infopost/ReportViewer.aspx?/InfoPost/{report_name}&pAssetNbr={asset_id}&rs:Format=CSV'
    r = s.get(oac_url, timeout=30)
    if r.status_code != 200:
        return [], {}
    oac_rows = list(csv.DictReader(io.StringIO(r.text)))
    
    # Also fetch location data for county/state
    loc_url = f'https://ebb.tceconnects.com/infopost/ReportViewer.aspx?/InfoPost/LocationDataDownload&assetNbr={asset_id}&rs:Format=CSV&rc:NoHeader=true'
    r2 = s.get(loc_url, timeout=30)
    loc_map = {}
    if r2.status_code == 200 and len(r2.text) > 100:
        loc_reader = csv.DictReader(io.StringIO(r2.text))
        for row in loc_reader:
            loc_id = row.get('Loc', '').strip()
            if loc_id:
                loc_map[loc_id] = {
                    'county': row.get('Loc Cnty', '').strip(),
                    'state': row.get('Loc St Abbrev', '').strip(),
                    'zone': row.get('Loc Zone', '').strip(),
                }
    
    return oac_rows, loc_map


# ============================================================
# TC PLUS (tcplus.com — JSON API, no auth)
# Covers: Great Lakes, GTN, Tuscarora
# IOC: POST /{pipeline}/IndexOfCustomers/Generate → JSON
# Unsub: POST /{pipeline}/Export/Generate → CSV (needs session fix)
# ============================================================

TCPLUS_PIPELINES = [
    {'path': 'Great%20Lakes', 'name': 'Great Lakes Gas Transmission Limited Partnership', 'short': 'Great Lakes'},
    {'path': 'GTN', 'name': 'Gas Transmission Northwest LLC', 'short': 'GTN'},
    {'path': 'Tuscarora', 'name': 'Tuscarora Gas Transmission Company', 'short': 'Tuscarora'},
]


def fetch_tcplus_ioc(pipeline_path):
    """Fetch Index of Customers from TC Plus platform. Returns JSON with contract details."""
    s = requests.Session()
    s.headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    s.headers['X-Requested-With'] = 'XMLHttpRequest'

    url = f'https://tcplus.com/{pipeline_path}/IndexOfCustomers/Generate'
    r = s.post(url, timeout=30)

    if r.status_code != 200:
        return {}

    try:
        data = r.json()
    except:
        return {}

    contracts = data.get('data', {}).get('ShipperGroup', [])

    cutoff = datetime.now() + timedelta(days=730)
    firm_mdq = 0
    expiring_2yr = 0
    num_contracts = 0
    shippers = set()

    for c in contracts:
        shipper = c.get('ShipperName', '').strip()
        rate = c.get('RateScheduleName', '')
        mdq_str = c.get('Mdq', '0')
        exp_date_str = c.get('ContractEndDate', '')

        if not shipper:
            continue

        try:
            mdq = int(str(mdq_str).replace(',', ''))
        except:
            mdq = 0

        if mdq == 0:
            continue

        num_contracts += 1
        shippers.add(shipper)

        # FT, FTS, FTS-1 etc. are firm transportation
        if 'FT' in rate.upper():
            firm_mdq += mdq

        if exp_date_str:
            try:
                ed = datetime.strptime(exp_date_str.strip()[:10], '%m/%d/%Y')
                if ed <= cutoff:
                    expiring_2yr += mdq
            except:
                pass

    return {
        'firm_mdq': firm_mdq,
        'expiring_2yr': expiring_2yr,
        'num_contracts': num_contracts,
        'num_shippers': len(shippers),
    }


# TODO: fetch_tcplus_unsub has a session/context bug — the Export/Generate
# endpoint returns empty or HTML instead of CSV. Needs investigation into
# whether a specific cookie or prior page visit sets the correct context.
# Uncomment and fix when ready.
#
# def fetch_tcplus_unsub(pipeline_path):
#     """Fetch Unsubscribed Capacity CSV from TC Plus Export endpoint."""
#     s = requests.Session()
#     s.headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
#
#     # First visit the unsub page to set session context for this pipeline
#     s.get(f'https://tcplus.com/{pipeline_path}/UnsubscribedCapacity', timeout=15)
#     time.sleep(1)
#
#     # Then export
#     url = f'https://tcplus.com/{pipeline_path}/Export/Generate'
#     form_data = {
#         'serviceTypeName': 'Ganesha.InfoPost.Service.UnsubscribedCapacity.UnsubscribedCapacityService, Ganesha.InfoPost.Service',
#         'filterTypeName': 'System.Int32',
#         'templateType': '7',
#         'exportType': '1',
#         'filter': '6',
#         'customExtension': '',
#     }
#
#     r = s.post(url, data=form_data, timeout=30)
#
#     if r.status_code != 200 or 'csv' not in r.headers.get('Content-Type', '').lower():
#         return []
#
#     # Parse CSV — skip header rows (first 5 lines are metadata)
#     lines = r.text.strip().split('\n')
#     data_lines = []
#     header_found = False
#
#     for line in lines:
#         if 'Location Name' in line:
#             header_found = True
#             continue
#         if header_found and line.strip():
#             reader = csv.reader(io.StringIO(line))
#             for row in reader:
#                 if len(row) >= 5:
#                     loc_name = row[0].strip().strip('"')
#                     loc_id = row[1].strip().strip('"')
#                     purp = row[2].strip().strip('"')
#                     unsub_str = row[4].strip().strip('"').replace(',', '')
#
#                     try:
#                         unsub = int(unsub_str)
#                     except:
#                         unsub = 0
#
#                     data_lines.append({
#                         'Loc': loc_id,
#                         'Loc_Name': loc_name,
#                         'Loc_Purp_Desc': purp,
#                         'Unsubscribed_Capacity': unsub,
#                     })
#
#     return data_lines


# ============================================================
# PARSING HELPERS
# ============================================================

def extract_hidden_fields(html):
    """Extract ASP.NET hidden form fields."""
    fields = {}
    for m in re.finditer(r'<input[^>]*type="hidden"[^>]*name="([^"]+)"[^>]*value="([^"]*)"', html):
        fields[m.group(1)] = m.group(2)
    return fields


def parse_km_xlsx(content):
    """Parse KM multi-row header XLSX capacity file."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    
    header_idx = None
    for i, row in enumerate(rows):
        if row and str(row[0]).strip() == 'Loc':
            header_idx = i
            break
    
    if header_idx is None:
        return []
    
    headers = [str(h).strip() if h else f'col{j}' for j, h in enumerate(rows[header_idx])]
    data = []
    for row in rows[header_idx + 1:]:
        if row and row[0] is not None:
            data.append(dict(zip(headers, row)))
    return data


def parse_williams_html(html):
    """Parse Williams OAC HTML report into structured data."""
    rows = re.findall(r'<TR[^>]*>(.*?)</TR>', html, re.DOTALL | re.I)
    data = []
    for row in rows:
        cells = re.findall(r'<TD[^>]*>(.*?)</TD>', row, re.DOTALL | re.I)
        clean = [re.sub(r'<[^>]+>', '', c).strip().replace('&nbsp;', '').replace('\n', ' ').strip() for c in cells]
        if len(clean) >= 10 and any(c.replace(',', '').replace('-', '').isdigit() for c in clean[6:10]):
            data.append({
                'Loc': clean[0],
                'Loc Purp Desc': clean[1],
                'Flow Ind': clean[2],
                'QTI': clean[3],
                'Loc Name': clean[4],
                'Loc Zn': clean[5],
                'Design Capacity': clean[6],
                'Operating Capacity': clean[7],
                'Total Scheduled Quantity': clean[8],
                'Operationally Available Capacity': clean[9],
            })
    return data


def parse_ioc_xlsx(content):
    """Parse KM IOC Excel into per-point contract aggregates."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    
    header_idx = None
    for i, row in enumerate(rows):
        if row and any('Shipper' in str(c) for c in row if c):
            header_idx = i
            break
    if header_idx is None:
        return {}
    
    headers = [str(h).strip() if h else f'col{j}' for j, h in enumerate(rows[header_idx])]
    data = [dict(zip(headers, row)) for row in rows[header_idx + 1:] if row and row[0] is not None]
    
    loc_caps = defaultdict(lambda: {'firm_mdq': 0, 'expiring_2yr': 0, 'num_contracts': 0, 'shippers': set()})
    
    cutoff = datetime.now() + timedelta(days=730)
    
    for d in data:
        point_id = str(d.get('Point ID', '')).strip()
        mdq = d.get('MDQ') or d.get('PT MDQ') or 0
        rate = str(d.get('Rate Sched', '')).strip()
        shipper = str(d.get('Shipper Name', '')).strip()
        exp_date = d.get('Contract Expiration Date', '')
        
        try:
            mdq = int(str(mdq).replace(',', '').strip())
        except:
            mdq = 0
        
        if not point_id or mdq == 0:
            continue
        
        loc_caps[point_id]['num_contracts'] += 1
        loc_caps[point_id]['shippers'].add(shipper)
        
        if 'FT' in rate.upper():
            loc_caps[point_id]['firm_mdq'] += mdq
        
        if exp_date:
            try:
                if isinstance(exp_date, datetime):
                    ed = exp_date
                else:
                    ed = datetime.strptime(str(exp_date).strip()[:10], '%m/%d/%Y')
                if ed <= cutoff:
                    loc_caps[point_id]['expiring_2yr'] += mdq
            except:
                pass
    
    result = {}
    for loc_id, info in loc_caps.items():
        result[loc_id] = {
            'firm_mdq': info['firm_mdq'],
            'expiring_2yr': info['expiring_2yr'],
            'num_contracts': info['num_contracts'],
            'num_shippers': len(info['shippers']),
        }
    return result


def parse_int_safe(val):
    """Parse integer from possibly comma-formatted string."""
    if val is None:
        return 0
    s = str(val).replace(',', '').strip()
    try:
        return int(float(s))
    except:
        return 0


# ============================================================
# MAIN PIPELINE
# ============================================================

def fetch_all_capacity():
    """Fetch capacity data from all platforms. Returns list of pipeline dicts."""
    pipelines = []
    
    # Kinder Morgan
    for pl in KM_PIPELINES:
        print(f"Fetching KM {pl['short']}...")
        try:
            caps = fetch_km_capacity(pl['code'])
            locs = fetch_km_locations(pl['code'])
            ioc = fetch_km_ioc(pl['code'])
            
            points = []
            for c in caps:
                loc_id = str(c.get('Loc', '')).strip()
                if not loc_id:
                    continue
                loc = locs.get(loc_id, {})
                purp = str(c.get('Loc Purp Desc', '')).strip()
                
                dc = parse_int_safe(c.get('Design Capacity'))
                sched = parse_int_safe(c.get('Total Scheduled Quantity'))
                avail_key = next((k for k in c.keys() if 'Operationally' in str(k)), None)
                avail = parse_int_safe(c.get(avail_key)) if avail_key else 0
                
                if dc == 0:
                    continue
                
                ptype = 'delivery' if 'Delivery' in purp else ('receipt' if 'Receipt' in purp else 'other')
                
                pt = {
                    'id': loc_id,
                    'name': str(c.get('Loc Name', '')).strip()[:50],
                    'type': ptype,
                    'county': str(loc.get('Loc Cnty', '')).strip(),
                    'state': str(loc.get('Loc St Abbrev', '')).strip(),
                    'design': dc,
                    'scheduled': sched,
                    'available': avail,
                    'utilization': round(sched / dc * 100) if dc > 0 else 0,
                    'connected': str(loc.get('Up/Dn Name', '')).strip()[:50] if loc.get('Up/Dn Name') else '',
                }
                
                # Add IOC data if available
                if loc_id in ioc:
                    pt['firm_contracted'] = ioc[loc_id]['firm_mdq']
                    pt['expiring_2yr'] = ioc[loc_id]['expiring_2yr']
                    pt['num_shippers'] = ioc[loc_id]['num_shippers']
                    pt['num_contracts'] = ioc[loc_id]['num_contracts']
                
                points.append(pt)
            
            pipelines.append({
                'name': pl['name'],
                'short': pl['short'],
                'updated': TODAY,
                'points': points,
            })
            print(f"  {pl['short']}: {len(points)} points, {len(ioc)} IOC locations")
        except Exception as e:
            print(f"  ERROR {pl['short']}: {e}")
        time.sleep(2)
    
    # Williams
    for pl in WILLIAMS_PIPELINES:
        print(f"Fetching Williams {pl['short']}...")
        try:
            caps = fetch_williams_capacity(pl['buid'])
            
            points = []
            for c in caps:
                if 'Segment' in c.get('Loc Purp Desc', ''):
                    continue
                
                dc = parse_int_safe(c.get('Design Capacity'))
                sched = parse_int_safe(c.get('Total Scheduled Quantity'))
                avail = parse_int_safe(c.get('Operationally Available Capacity'))
                
                if dc == 0:
                    continue
                
                flow = c.get('Flow Ind', '')
                ptype = 'delivery' if 'D' in flow else ('receipt' if 'R' in flow else 'other')
                
                points.append({
                    'id': c.get('Loc', ''),
                    'name': c.get('Loc Name', '').strip()[:50],
                    'type': ptype,
                    'county': '',
                    'state': '',
                    'zone': c.get('Loc Zn', ''),
                    'design': dc,
                    'scheduled': sched,
                    'available': avail,
                    'utilization': round(sched / dc * 100) if dc > 0 else 0,
                    'connected': '',
                })
            
            pipelines.append({
                'name': pl['name'],
                'short': pl['short'],
                'updated': TODAY,
                'points': points,
            })
            print(f"  {pl['short']}: {len(points)} points")
        except Exception as e:
            print(f"  ERROR {pl['short']}: {e}")
        time.sleep(2)
    
    # Enbridge
    for pl in ENBRIDGE_PIPELINES:
        print(f"Fetching Enbridge {pl['short']}...")
        try:
            caps = fetch_enbridge_capacity(pl['bu'])
            ioc = fetch_enbridge_ioc(pl['bu'])

            points = []
            for c in caps:
                if 'Segment' in c.get('Loc_Purp_Desc', ''):
                    continue

                dc = parse_int_safe(c.get('Total_Design_Capacity'))
                sched = parse_int_safe(c.get('Total_Scheduled_Quantity'))
                avail = parse_int_safe(c.get('Operationally_Available_Capacity'))

                if dc == 0:
                    continue

                flow = c.get('Flow_Ind_Desc', '')
                ptype = 'delivery' if 'Delivery' in flow else ('receipt' if 'Receipt' in flow else 'other')

                pt = {
                    'id': c.get('Loc', ''),
                    'name': c.get('Loc_Name', '').strip()[:50],
                    'type': ptype,
                    'county': '',
                    'state': '',
                    'zone': c.get('Loc_Zn', ''),
                    'design': dc,
                    'scheduled': sched,
                    'available': avail,
                    'utilization': round(sched / dc * 100) if dc > 0 else 0,
                    'connected': '',
                }

                # Apply pipeline-level IOC stats to all points (same pattern as ET)
                if ioc:
                    pt['firm_contracted'] = ioc.get('firm_mdq', 0)
                    pt['expiring_2yr'] = ioc.get('expiring_2yr', 0)
                    pt['num_shippers'] = ioc.get('num_shippers', 0)
                    pt['num_contracts'] = ioc.get('num_contracts', 0)

                points.append(pt)

            pipelines.append({
                'name': pl['name'],
                'short': pl['short'],
                'updated': TODAY,
                'points': points,
            })
            print(f"  {pl['short']}: {len(points)} points, {ioc.get('num_contracts', 0)} IOC contracts")
        except Exception as e:
            print(f"  ERROR {pl['short']}: {e}")
        time.sleep(2)
    
    # Energy Transfer
    for pl in ET_PIPELINES:
        print(f"Fetching ET {pl['short']}...")
        try:
            caps = fetch_et_capacity(pl['base'], pl['asset'])
            ioc = fetch_et_ioc(pl['base'], pl['asset'])
            
            points = []
            for c in caps:
                if 'Segment' in c.get('Loc_Purp_Desc', ''):
                    continue
                
                dc = parse_int_safe(c.get('Total_Design_Capacity'))
                sched = parse_int_safe(c.get('Total_Scheduled_Quantity'))
                avail = parse_int_safe(c.get('Operationally_Available_Capacity'))
                
                if dc == 0:
                    continue
                
                purp = c.get('Loc_Purp_Desc', '')
                ptype = 'delivery' if 'Delivery' in purp else ('receipt' if 'Receipt' in purp else 'other')
                
                pt = {
                    'id': c.get('Loc', ''),
                    'name': c.get('Loc_Name', '').strip()[:50],
                    'type': ptype,
                    'county': c.get('County', '').replace(' County', '').strip(),
                    'state': c.get('State', '').strip(),
                    'design': dc,
                    'scheduled': sched,
                    'available': avail,
                    'utilization': round(sched / dc * 100) if dc > 0 else 0,
                    'connected': c.get('Operator', '').strip()[:50],
                }
                
                # Add IOC data if available (ET IOC is by shipper, not by point)
                # Apply pipeline-level IOC stats to all points
                if ioc:
                    pt['firm_contracted'] = ioc.get('firm_mdq', 0)
                    pt['expiring_2yr'] = ioc.get('expiring_2yr', 0)
                    pt['num_shippers'] = ioc.get('num_shippers', 0)
                    pt['num_contracts'] = ioc.get('num_contracts', 0)
                
                points.append(pt)
            
            pipelines.append({
                'name': pl['name'],
                'short': pl['short'],
                'updated': TODAY,
                'points': points,
            })
            print(f"  {pl['short']}: {len(points)} points, {ioc.get('num_contracts', 0)} IOC contracts")
        except Exception as e:
            print(f"  ERROR {pl['short']}: {e}")
        time.sleep(2)
    
    # CenterPoint / EGT (Energy Transfer CSV downloads)
    for pl in EGT_PIPELINES:
        print(f"Fetching EGT {pl['short']}...")
        try:
            oac_rows, loc_map = fetch_egt_capacity(pl['asset'])

            points = []
            for c in oac_rows:
                loc_id = c.get('LOCATION', '').strip()
                if not loc_id:
                    continue

                purp = c.get('LOC PURPOSE', '').strip()
                if 'Segment' in purp:
                    continue

                dc = parse_int_safe(c.get('DESIGN CAPACITY'))
                sched = parse_int_safe(c.get('SCHED QTY'))
                avail = parse_int_safe(c.get('AVAILABLE CAPACITY'))

                if dc == 0:
                    continue

                flow = c.get('FLOW DIR', '').strip()
                ptype = 'delivery' if 'Delivery' in flow else ('receipt' if 'Receipt' in flow else 'other')
                loc_info = loc_map.get(loc_id, {})

                points.append({
                    'id': loc_id,
                    'name': c.get('LOCATION NAME', '').strip()[:50],
                    'type': ptype,
                    'county': loc_info.get('county', '').replace(' County', '').strip(),
                    'state': loc_info.get('state', '').strip(),
                    'design': dc,
                    'scheduled': sched,
                    'available': avail,
                    'utilization': round(sched / dc * 100) if dc > 0 else 0,
                    'connected': loc_info.get('operator', '').strip()[:50],
                })

            pipelines.append({
                'name': pl['name'],
                'short': pl['short'],
                'updated': TODAY,
                'points': points,
            })
            print(f"  {pl['short']}: {len(points)} points (from {len(oac_rows)} rows, {len(loc_map)} locations)")
        except Exception as e:
            print(f"  ERROR {pl['short']}: {e}")
        time.sleep(2)

    # TC Energy (eConnects)
    for pl in TC_PIPELINES:
        print(f"Fetching TC {pl['short']}...")
        try:
            oac_rows, loc_map = fetch_tc_capacity(pl['asset'], pl['report'])
            
            # Aggregate by location (may have multiple rows per location for rate schedules)
            loc_data = {}
            for row in oac_rows:
                loc_id = row.get('Location', '').strip()
                if not loc_id:
                    continue
                
                purp = row.get('LocPurpDesc', '').strip()
                if 'Segment' in purp:
                    continue
                
                dc = parse_int_safe(row.get('DesignCapacity'))
                sched = parse_int_safe(row.get('TotalSchedQty', row.get('TotalScheduledQuantity', '')))
                avail = parse_int_safe(row.get('OperationallyAvailableCapacity'))
                
                if dc == 0:
                    continue
                
                key = f"{loc_id}|{purp}"
                if key not in loc_data or dc > loc_data[key]['design']:
                    ptype = 'delivery' if 'Delivery' in purp else ('receipt' if 'Receipt' in purp else 'other')
                    loc_info = loc_map.get(loc_id, {})
                    
                    loc_data[key] = {
                        'id': loc_id,
                        'name': row.get('LocationName', '').strip()[:50],
                        'type': ptype,
                        'county': loc_info.get('county', ''),
                        'state': loc_info.get('state', ''),
                        'design': dc,
                        'scheduled': sched,
                        'available': avail,
                        'utilization': round(sched / dc * 100) if dc > 0 else 0,
                        'connected': '',
                    }
            
            points = list(loc_data.values())
            pipelines.append({
                'name': pl['name'],
                'short': pl['short'],
                'updated': TODAY,
                'points': points,
            })
            print(f"  {pl['short']}: {len(points)} points (from {len(oac_rows)} rows, {len(loc_map)} locations)")
        except Exception as e:
            print(f"  ERROR {pl['short']}: {e}")
        time.sleep(2)

    # TC Plus (IOC only — no capacity fetch yet, unsub endpoint has session bug)
    for pl in TCPLUS_PIPELINES:
        print(f"Fetching TC Plus {pl['short']}...")
        try:
            ioc = fetch_tcplus_ioc(pl['path'])

            # TC Plus pipelines: IOC-only for now (no OAC endpoint discovered)
            # Create a single pipeline-level entry so IOC data is tracked
            if ioc and ioc.get('num_contracts', 0) > 0:
                pipelines.append({
                    'name': pl['name'],
                    'short': pl['short'],
                    'updated': TODAY,
                    'points': [],
                    'ioc_totals': ioc,
                })
                print(f"  {pl['short']}: {ioc.get('num_contracts', 0)} IOC contracts, {ioc.get('num_shippers', 0)} shippers, {ioc.get('firm_mdq', 0):,} firm MDQ")
            else:
                print(f"  {pl['short']}: no IOC data returned")
        except Exception as e:
            print(f"  ERROR {pl['short']}: {e}")
        time.sleep(2)

    return pipelines


def update_history(pipelines):
    """Append today's utilization snapshot to history file for rolling averages."""
    history = {}
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE) as f:
            history = json.load(f)
    
    for pl in pipelines:
        for pt in pl['points']:
            key = f"{pl['short']}|{pt['id']}"
            if key not in history:
                history[key] = {'snapshots': []}
            
            history[key]['snapshots'].append({
                'date': TODAY,
                'scheduled': pt['scheduled'],
                'available': pt['available'],
                'utilization': pt['utilization'],
            })
            
            # Keep only last 45 days
            cutoff = (datetime.now() - timedelta(days=45)).strftime('%Y-%m-%d')
            history[key]['snapshots'] = [
                s for s in history[key]['snapshots'] if s['date'] >= cutoff
            ]
    
    with open(HISTORY_FILE, 'w') as f:
        json.dump(history, f)
    
    print(f"History updated: {len(history)} tracked points")
    return history


def compute_rolling_stats(pipelines, history):
    """Add 30-day rolling average and peak to each point."""
    for pl in pipelines:
        for pt in pl['points']:
            key = f"{pl['short']}|{pt['id']}"
            snaps = history.get(key, {}).get('snapshots', [])
            
            # Filter to last 30 days
            cutoff = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
            recent = [s for s in snaps if s['date'] >= cutoff]
            
            if len(recent) >= 2:
                utils = [s['utilization'] for s in recent]
                scheds = [s['scheduled'] for s in recent]
                pt['avg_utilization_30d'] = round(sum(utils) / len(utils), 1)
                pt['peak_utilization_30d'] = max(utils)
                pt['avg_scheduled_30d'] = round(sum(scheds) / len(scheds))
                pt['peak_scheduled_30d'] = max(scheds)
                pt['days_of_data'] = len(recent)


# ============================================================
# ENTRY POINT
# ============================================================

# ============================================================
# GEOCODING + HIFLD MATCHING
# ============================================================

PIPELINE_HIFLD_MAP = {
    'El Paso Natural Gas Company': ['EL PASO NATURAL GAS COMPANY'],
    'Tennessee Gas Pipeline Company': ['TENNESSEE GAS PIPELINE'],
    'Natural Gas Pipeline Company of America': ['NATURAL GAS PIPELINE (KINDER MORGAN)'],
    'Midcontinent Express Pipeline': ['MIDCONTINENT EXPRESS PIPELINE'],
    'Southern Natural Gas Company, LLC': ['SOUTHERN NATURAL GAS COMPANY LLC'],
    'Transcontinental Gas Pipe Line Company (Transco)': ['TRANSCONTINENTAL GAS PIPE LINE COMPANY, LLC'],
    'Texas Eastern Transmission, LP': ['TEXAS EASTERN TRANSMISSION'],
    'Florida Gas Transmission Company, LLC': ['FLORIDA GAS TRANSMISSION COMPANY'],
    'Colorado Interstate Gas Company, LLC': ['COLORADO INTERSTATE GAS COMPANY'],
    'Algonquin Gas Transmission, LLC': ['ALGONQUIN GAS TRANSMISSION'],
    'Maritimes & Northeast Pipeline, LLC': ['MARITIMES AND NORTHEAST PIPELINE'],
    'East Tennessee Natural Gas, LLC': ['EAST TENNESSEE NATURAL GAS'],
    'Panhandle Eastern Pipe Line Company': ['PANHANDLE EASTERN PIPE LINE COMPANY'],
    'Trunkline Gas Company, LLC': ['TRUNKLINE LNG COMPANY', 'TRUNKLINE GAS COMPANY'],
    'Rover Pipeline LLC': ['ROVER PIPELINE'],
    'ANR Pipeline Company': ['ANR PIPELINE COMPANY'],
    'Columbia Gas Transmission, LLC': ['COLUMBIA GAS TRANSMISSION'],
    'Columbia Gulf Transmission, LLC': ['COLUMBIA GULF TRANSMISSION'],
    'Northern Border Pipeline Company': ['NORTHERN BORDER PIPELINE COMPANY'],
    'CenterPoint Energy Gas Transmission Company': ['CENTERPOINT ENERGY', 'ENABLE GAS TRANSMISSION'],
    'Great Lakes Gas Transmission Limited Partnership': ['GREAT LAKES GAS TRANS LTD'],
    'Gas Transmission Northwest LLC': ['GAS TRANSMISSION NORTHWEST'],
    'Tuscarora Gas Transmission Company': ['TUSCARORA GAS TRANSMISSION COMPANY'],
}

ZONE_COORDS = {
    # Transco zones
    'Transco': {
        '1': {'lat': 29.5, 'lng': -96.5, 'state': 'TX'},
        '2': {'lat': 30.5, 'lng': -91.0, 'state': 'LA'},
        '3': {'lat': 32.0, 'lng': -88.5, 'state': 'MS/AL'},
        '4': {'lat': 33.5, 'lng': -84.5, 'state': 'GA'},
        '5': {'lat': 36.5, 'lng': -79.0, 'state': 'NC/VA'},
        '6': {'lat': 40.5, 'lng': -74.5, 'state': 'NJ/NY'},
    },
    # Texas Eastern zones
    'Texas Eastern': {
        '1': {'lat': 29.8, 'lng': -94.5, 'state': 'TX'},
        'STX': {'lat': 28.5, 'lng': -97.5, 'state': 'TX'},
        '2': {'lat': 31.0, 'lng': -91.5, 'state': 'LA'},
        '3': {'lat': 35.5, 'lng': -86.0, 'state': 'TN/KY'},
        'ELA': {'lat': 30.5, 'lng': -90.0, 'state': 'LA'},
        'WLA': {'lat': 30.0, 'lng': -93.0, 'state': 'LA'},
        'ETX': {'lat': 31.5, 'lng': -94.5, 'state': 'TX'},
        'SLA': {'lat': 29.5, 'lng': -91.0, 'state': 'LA'},
        'M1': {'lat': 40.3, 'lng': -75.0, 'state': 'PA/NJ'},
        'M2': {'lat': 40.8, 'lng': -74.0, 'state': 'NJ'},
        'M3': {'lat': 41.2, 'lng': -73.0, 'state': 'CT'},
    },
    # Algonquin zones
    'Algonquin': {
        'SE': {'lat': 41.0, 'lng': -73.5, 'state': 'CT'},
        'NE': {'lat': 42.0, 'lng': -72.0, 'state': 'MA'},
        'AG': {'lat': 41.5, 'lng': -73.0, 'state': 'CT/NY'},
    },
    # East Tennessee zones
    'East Tennessee': {
        '100': {'lat': 37.0, 'lng': -81.0, 'state': 'VA'},
        '200': {'lat': 36.0, 'lng': -83.0, 'state': 'TN'},
        '300': {'lat': 35.5, 'lng': -84.5, 'state': 'TN'},
        '400': {'lat': 34.5, 'lng': -84.0, 'state': 'GA'},
        'ET': {'lat': 36.0, 'lng': -83.0, 'state': 'TN'},
    },
}


def load_hifld_points():
    """Load or fetch HIFLD gas receipt/delivery points."""
    if os.path.exists(HIFLD_CACHE):
        with open(HIFLD_CACHE) as f:
            return json.load(f)
    
    print("Fetching HIFLD gas points (first run)...")
    base_url = "https://services5.arcgis.com/HDRa0B57OVrv2E1q/arcgis/rest/services/Natural_Gas_Receipt_Delivery_Points/FeatureServer/0/query"
    all_pts = []
    offset = 0
    while True:
        r = requests.get(base_url, params={
            'where': "COUNTRY='USA'",
            'outFields': 'NAME,STATE,COUNTY,TYPE,COMPNAME,LATITUDE,LONGITUDE',
            'resultOffset': offset, 'resultRecordCount': 2000, 'f': 'json'
        }, headers={'user-agent': 'Mozilla/5.0'})
        features = r.json().get('features', [])
        for f in features:
            a = f['attributes']
            all_pts.append({
                'name': a.get('NAME', ''),
                'state': a.get('STATE', ''),
                'county': a.get('COUNTY', ''),
                'company': a.get('COMPNAME', ''),
                'lat': a.get('LATITUDE', 0),
                'lng': a.get('LONGITUDE', 0),
            })
        if len(features) < 2000:
            break
        offset += 2000
        time.sleep(1)
    
    with open(HIFLD_CACHE, 'w') as f:
        json.dump(all_pts, f)
    print(f"  Cached {len(all_pts)} HIFLD points")
    return all_pts


def normalize_name(name):
    """Normalize point name for fuzzy matching."""
    n = name.upper().strip()
    for prefix in ['EPNG/', 'SNG/', 'TGP/', 'NGPL/', 'TETCO/', 'GS-', 'GS ', 'KMTP/', 'MEP/']:
        if n.startswith(prefix):
            n = n[len(prefix):]
    n = re.sub(r'\([^)]*\)', '', n)
    for word in [' DEL', ' REC', ' DELIVERY', ' RECEIPT', ' METER', ' STATION', ' STA',
                 ' SHIPPER', ' DEDUCT', ' DED', ' PLANT', ' POWER',
                 ' LLC', ' INC', ' CORP', ' CO', ' COMPANY']:
        n = n.replace(word, '')
    return re.sub(r'\s+', ' ', n).strip()


def name_match_score(n1, n2):
    """Score similarity between two point names."""
    from difflib import SequenceMatcher
    a, b = normalize_name(n1), normalize_name(n2)
    if a == b:
        return 1.0
    if a in b or b in a:
        return 0.9
    wa, wb = a.split(), b.split()
    if wa and wb and wa[0] == wb[0] and len(wa[0]) > 3:
        return max(0.6, SequenceMatcher(None, a, b).ratio())
    return SequenceMatcher(None, a, b).ratio()


def geocode_counties(counties_needing_coords):
    """Batch geocode county names via Geocodio."""
    if not counties_needing_coords or not GEOCODIO_KEY:
        return {}
    
    queries = [f"{nc.split('|')[0]} County, {nc.split('|')[1]}" for nc in counties_needing_coords]
    results = {}
    
    # Geocodio batch limit is 10,000
    for i in range(0, len(queries), 500):
        batch = queries[i:i+500]
        batch_keys = counties_needing_coords[i:i+500]
        try:
            r = requests.post(f"https://api.geocod.io/v1.7/geocode?api_key={GEOCODIO_KEY}", json=batch)
            if r.status_code == 200:
                for j, result in enumerate(r.json().get('results', [])):
                    locs = result.get('response', {}).get('results', [])
                    if locs:
                        results[batch_keys[j]] = {
                            'lat': locs[0]['location']['lat'],
                            'lng': locs[0]['location']['lng'],
                        }
        except Exception as e:
            print(f"  Geocodio error: {e}")
        time.sleep(1)
    
    return results


def geocode_and_locate(pipelines):
    """Add lat/lng to all points using HIFLD matching, county geocoding, and zone fallback."""
    import random
    random.seed(42)
    
    # Load caches
    county_coords = {}
    if os.path.exists(COUNTY_CACHE):
        with open(COUNTY_CACHE) as f:
            county_coords = json.load(f)
    
    hifld = load_hifld_points()
    
    # Build HIFLD lookups
    hifld_by_company = defaultdict(list)
    hifld_by_company_county = defaultdict(list)
    for h in hifld:
        comp = h['company'].upper()
        hifld_by_company[comp].append(h)
        county = h.get('county', '').upper()
        if county and county != 'NOT AVAILABLE':
            hifld_by_company_county[f"{comp}|{county}|{h['state']}"].append(h)
    
    # Find new counties that need geocoding
    new_counties = []
    for pl in pipelines:
        for pt in pl['points']:
            county, state = pt.get('county', ''), pt.get('state', '')
            if county and state:
                key = f"{county}|{state}"
                if key not in county_coords:
                    new_counties.append(key)
    
    new_counties = list(set(new_counties))
    if new_counties:
        print(f"\nGeocoding {len(new_counties)} new counties...")
        new_coords = geocode_counties(new_counties)
        county_coords.update(new_coords)
        with open(COUNTY_CACHE, 'w') as f:
            json.dump(county_coords, f)
        print(f"  Geocoded {len(new_coords)}/{len(new_counties)}")
    
    # Match each point
    stats = {'hifld': 0, 'county': 0, 'zone': 0, 'total': 0}
    
    for pl in pipelines:
        hifld_companies = PIPELINE_HIFLD_MAP.get(pl['name'], [])
        zone_map = ZONE_COORDS.get(pl['short'], {})
        
        for pt in pl['points']:
            stats['total'] += 1
            county = pt.get('county', '').upper()
            state = pt.get('state', '')
            matched = False
            
            # Pass 1: HIFLD name match (within state if available, global if not)
            best_score = 0
            best_match = None
            
            for hc in hifld_companies:
                if state and '/' not in state:
                    candidates = [h for h in hifld_by_company.get(hc, []) if h['state'] == state]
                elif state and '/' in state:
                    states = state.split('/')
                    candidates = [h for h in hifld_by_company.get(hc, []) if h['state'] in states]
                else:
                    candidates = hifld_by_company.get(hc, [])
                
                for c in candidates:
                    score = name_match_score(pt['name'], c['name'])
                    if county and c.get('county', '').upper() not in ('', 'NOT AVAILABLE'):
                        if county in c['county'].upper() or c['county'].upper() in county:
                            score += 0.1
                    if score > best_score:
                        best_score = score
                        best_match = c
            
            # Pass 2: County-only match (single point of same type in county)
            if best_score < 0.4 and county and state:
                ptype_map = {'delivery': 'DELIVERY', 'receipt': 'RECEIPT'}
                type_kw = ptype_map.get(pt['type'], '')
                for hc in hifld_companies:
                    key = f"{hc}|{county}|{state}"
                    cands = hifld_by_company_county.get(key, [])
                    typed = [c for c in cands if type_kw in c.get('type', '').upper()] if type_kw else cands
                    if len(typed) == 1:
                        best_score = 0.7
                        best_match = typed[0]
                        break
            
            # Apply HIFLD match
            if best_score >= 0.4 and best_match:
                pt['lat'] = round(best_match['lat'], 5)
                pt['lng'] = round(best_match['lng'], 5)
                pt['loc_accuracy'] = 'hifld'
                if best_match.get('county', 'NOT AVAILABLE') != 'NOT AVAILABLE' and not pt.get('county'):
                    pt['county'] = best_match['county'].title()
                if best_match.get('state') and not pt.get('state'):
                    pt['state'] = best_match['state']
                stats['hifld'] += 1
                matched = True
            
            # Fallback: county centroid
            if not matched and county and state:
                key = f"{pt['county']}|{state}"
                if key in county_coords:
                    cc = county_coords[key]
                    pt['lat'] = round(cc['lat'] + random.uniform(-0.05, 0.05), 5)
                    pt['lng'] = round(cc['lng'] + random.uniform(-0.05, 0.05), 5)
                    pt['loc_accuracy'] = 'county'
                    stats['county'] += 1
                    matched = True
            
            # Fallback: zone centroid
            if not matched:
                zone = pt.get('zone', '')
                zc = zone_map.get(zone)
                if zc:
                    pt['lat'] = round(zc['lat'] + random.uniform(-0.8, 0.8), 5)
                    pt['lng'] = round(zc['lng'] + random.uniform(-0.8, 0.8), 5)
                    pt['state'] = pt.get('state') or zc.get('state', '')
                    pt['loc_accuracy'] = 'zone'
                    stats['zone'] += 1
                    matched = True
            
            if not matched:
                pt['loc_accuracy'] = 'zone'
                stats['zone'] += 1
        
    # Remove points with no coordinates
    for pl in pipelines:
        pl['points'] = [p for p in pl['points'] if 'lat' in p]
    
    total = stats['total']
    print(f"\nGeocoding results:")
    print(f"  HIFLD precise: {stats['hifld']} ({stats['hifld']*100//max(total,1)}%)")
    print(f"  County centroid: {stats['county']} ({stats['county']*100//max(total,1)}%)")
    print(f"  Zone estimate: {stats['zone']} ({stats['zone']*100//max(total,1)}%)")

if __name__ == '__main__':
    print(f"=== Gas Interconnect Refresh: {TODAY} ===\n")
    
    # Fetch fresh data
    pipelines = fetch_all_capacity()
    
    total_pts = sum(len(p['points']) for p in pipelines)
    print(f"\nTotal: {len(pipelines)} pipelines, {total_pts} points")
    
    # Update history for rolling averages
    history = update_history(pipelines)
    
    # Compute rolling stats
    compute_rolling_stats(pipelines, history)
    
    # Geocode + HIFLD coordinate matching
    geocode_and_locate(pipelines)
    
    output = {'pipelines': pipelines}
    with open(OUTPUT_FILE, 'w') as f:
        json.dump(output, f)
    
    print(f"\nOutput: {os.path.getsize(OUTPUT_FILE) / 1024:.0f} KB")
    print("Done.")


