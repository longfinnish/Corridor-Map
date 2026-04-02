#!/usr/bin/env python3
"""
Build utility_territories GeoJSON + PMTiles from HIFLD polygons + EIA 861 2024 data.
Steps: 1) Fetch HIFLD polygons, 2) Parse EIA Excel, 3) Join, 4) Write GeoJSON, 5) Convert to PMTiles
"""
import json, os, sys, time, math, struct, io, hashlib, gzip
import requests
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(os.path.dirname(SCRIPT_DIR), 'data')
OUT_GEOJSON = os.path.join(DATA_DIR, 'utility_territories.geojson')
OUT_PMTILES = os.path.join(DATA_DIR, 'utility_territories.pmtiles')

HIFLD_URL = 'https://services3.arcgis.com/OYP7N6mAJJCyH6hd/arcgis/rest/services/Electric_Retail_Service_Territories_HIFLD/FeatureServer/0/query'
HIFLD_FIELDS = 'ID,NAME,STATE,TYPE,CNTRL_AREA,PLAN_AREA,HOLDING_CO,WEBSITE,TELEPHONE,REGULATED'

# ── Step 1: Fetch HIFLD polygons ──
def fetch_hifld():
    print('[1/5] Fetching HIFLD utility territory polygons...')
    # Get count first
    r = requests.get(HIFLD_URL, params={
        'where': '1=1', 'returnCountOnly': 'true', 'f': 'json'
    }, timeout=30)
    total = r.json().get('count', 0)
    print(f'  Total records: {total}')

    features = []
    offset = 0
    batch = 250  # Small batches — polygon geometry is heavy
    while offset < total:
        print(f'  Fetching {offset}–{offset+batch} of {total}...')
        params = {
            'where': '1=1',
            'outFields': HIFLD_FIELDS,
            'outSR': '4326',
            'f': 'geojson',
            'resultOffset': offset,
            'resultRecordCount': batch
        }
        for attempt in range(3):
            try:
                resp = requests.get(HIFLD_URL, params=params, timeout=180)
                resp.raise_for_status()
                break
            except Exception as e:
                if attempt < 2:
                    print(f'    Retry {attempt+1}: {e}')
                    time.sleep(5 * (attempt + 1))
                else:
                    raise
        data = resp.json()
        batch_features = data.get('features', [])
        if not batch_features:
            break
        features.extend(batch_features)
        offset += batch
        time.sleep(1)

    print(f'  Fetched {len(features)} polygons')
    return features


# ── Step 2: Parse EIA 861 2024 Excel files ──
def parse_eia_value(val):
    """Parse EIA value: '.' → None, empty → None, number strings → float"""
    if val is None or val == '' or val == '.':
        return None
    try:
        v = float(val)
        if v == -999999:
            return None
        return v
    except (ValueError, TypeError):
        return None

def parse_operational_data():
    print('[2a/5] Parsing Operational_Data_2024.xlsx...')
    path = os.path.join(SCRIPT_DIR, 'Operational_Data_2024.xlsx')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['States']
    ops = {}  # utility_id → {fields}
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
        uid = row[1]  # Column B = Utility Number
        if uid is None:
            continue
        uid = int(uid)
        # For multi-state utilities, keep the first (or aggregate peaks by max)
        summer = parse_eia_value(row[6])   # G
        winter = parse_eia_value(row[7])   # H
        net_gen = parse_eia_value(row[8])  # I
        purchases = parse_eia_value(row[9]) # J
        ownership = str(row[4]) if row[4] else None  # E
        nerc = str(row[5]) if row[5] else None  # F

        if uid not in ops:
            ops[uid] = {
                'summer_peak_mw': summer,
                'winter_peak_mw': winter,
                'net_generation_mwh': net_gen,
                'total_purchases_mwh': purchases,
                'ownership_eia': ownership,
                'nerc_region': nerc
            }
        else:
            # Aggregate across states: sum generation/purchases, max peaks
            existing = ops[uid]
            if summer is not None:
                existing['summer_peak_mw'] = max(existing['summer_peak_mw'] or 0, summer)
            if winter is not None:
                existing['winter_peak_mw'] = max(existing['winter_peak_mw'] or 0, winter)
            if net_gen is not None:
                existing['net_generation_mwh'] = (existing['net_generation_mwh'] or 0) + net_gen
            if purchases is not None:
                existing['total_purchases_mwh'] = (existing['total_purchases_mwh'] or 0) + purchases
    wb.close()
    print(f'  Parsed {len(ops)} utilities from Operational Data')
    return ops

def parse_sales_data():
    print('[2b/5] Parsing Sales_Ult_Cust_2024.xlsx...')
    path = os.path.join(SCRIPT_DIR, 'Sales_Ult_Cust_2024.xlsx')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['States']
    customers = {}  # utility_id → total_customers
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
        uid = row[1]  # Column B
        if uid is None:
            continue
        uid = int(uid)
        # Column L (11) = Residential, O (14) = Commercial, R (17) = Industrial, U (20) = Transportation
        res = parse_eia_value(row[11])
        com = parse_eia_value(row[14])
        ind = parse_eia_value(row[17])
        trans = parse_eia_value(row[20]) if len(row) > 20 else None
        total = sum(v for v in [res, com, ind, trans] if v is not None)
        customers[uid] = customers.get(uid, 0) + total
    wb.close()
    print(f'  Parsed customer counts for {len(customers)} utilities')
    return customers


# ── Step 3: Join ──
def join_data(features, ops, customers):
    print('[3/5] Joining EIA data to HIFLD polygons...')
    matched = 0
    for f in features:
        p = f.get('properties', {})
        uid = p.get('ID')
        if uid is not None:
            try:
                uid = int(uid)
            except (ValueError, TypeError):
                uid = None

        # Clean HIFLD sentinel values
        for key in list(p.keys()):
            if p[key] == -999999 or p[key] == '-999999':
                p[key] = None

        # Build clean properties
        new_props = {
            'utility_id': uid,
            'name': p.get('NAME'),
            'state': p.get('STATE'),
            'type': p.get('TYPE'),
            'holding_company': p.get('HOLDING_CO'),
            'control_area': p.get('CNTRL_AREA'),
            'regulated': p.get('REGULATED'),
            'website': p.get('WEBSITE'),
            'telephone': p.get('TELEPHONE'),
        }

        # Join EIA operational data
        if uid and uid in ops:
            matched += 1
            op = ops[uid]
            new_props['nerc_region'] = op['nerc_region']
            new_props['summer_peak_mw'] = op['summer_peak_mw']
            new_props['winter_peak_mw'] = op['winter_peak_mw']
            new_props['net_generation_mwh'] = op['net_generation_mwh']
            new_props['total_purchases_mwh'] = op['total_purchases_mwh']
            new_props['data_year'] = '2024'
        else:
            new_props['nerc_region'] = None
            new_props['summer_peak_mw'] = None
            new_props['winter_peak_mw'] = None
            new_props['net_generation_mwh'] = None
            new_props['total_purchases_mwh'] = None
            new_props['data_year'] = '2022'

        # Join customer counts
        if uid and uid in customers:
            new_props['total_customers'] = int(customers[uid])
        else:
            new_props['total_customers'] = None

        f['properties'] = new_props

    print(f'  Matched {matched}/{len(features)} polygons with EIA operational data')
    return features


# ── Step 4: Write GeoJSON ──
def write_geojson(features):
    print('[4/5] Writing GeoJSON...')
    geojson = {
        'type': 'FeatureCollection',
        'features': features
    }
    with open(OUT_GEOJSON, 'w') as f:
        json.dump(geojson, f)
    size_mb = os.path.getsize(OUT_GEOJSON) / 1024 / 1024
    print(f'  Wrote {OUT_GEOJSON} ({size_mb:.1f} MB, {len(features)} features)')
    return geojson


# ── Step 5: Convert to PMTiles ──
def build_pmtiles(geojson):
    """Convert GeoJSON polygons to PMTiles using mapbox-vector-tile + pmtiles packages."""
    print('[5/5] Converting to PMTiles...')
    try:
        import mapbox_vector_tile as mvt
        from shapely.geometry import shape, mapping, box
        from shapely.ops import transform
        from pmtiles.writer import Writer as PMWriter
        from pmtiles.tile import TileType, Compression, zxy_to_tileid
    except ImportError as e:
        print(f'  ERROR: Missing package: {e}')
        print(f'  GeoJSON saved at {OUT_GEOJSON} — convert with tippecanoe manually')
        return

    features = geojson['features']
    # Pre-parse geometries with shapely
    print('  Parsing geometries...')
    parsed = []
    for f in features:
        try:
            geom = shape(f['geometry'])
            if geom.is_valid and not geom.is_empty:
                parsed.append((geom, f['properties']))
        except Exception:
            continue
    print(f'  {len(parsed)} valid geometries')

    # Build spatial index (simple grid) for tile queries
    # We'll generate tiles at zoom levels 0-12
    min_zoom = 0
    max_zoom = 10

    def lng_lat_to_tile(lng, lat, zoom):
        n = 2 ** zoom
        x = int((lng + 180) / 360 * n)
        lat_rad = math.radians(lat)
        y = int((1 - math.log(math.tan(lat_rad) + 1/math.cos(lat_rad)) / math.pi) / 2 * n)
        x = max(0, min(n-1, x))
        y = max(0, min(n-1, y))
        return x, y

    def tile_bounds(x, y, z):
        n = 2 ** z
        lng1 = x / n * 360 - 180
        lng2 = (x + 1) / n * 360 - 180
        lat1 = math.degrees(math.atan(math.sinh(math.pi * (1 - 2 * (y + 1) / n))))
        lat2 = math.degrees(math.atan(math.sinh(math.pi * (1 - 2 * y / n))))
        return lng1, lat1, lng2, lat2

    def geom_to_tile_coords(geom, tile_bbox, extent=4096):
        """Transform geometry from lng/lat to tile pixel coordinates."""
        lng1, lat1, lng2, lat2 = tile_bbox
        dx = lng2 - lng1
        dy = lat2 - lat1
        if dx == 0 or dy == 0:
            return None

        def transform_coord(x, y):
            px = (x - lng1) / dx * extent
            py = (lat2 - y) / dy * extent  # flip y
            return (round(px), round(py))

        try:
            from shapely.ops import transform as shapely_transform
            from shapely.geometry import mapping as shapely_mapping
            transformed = shapely_transform(lambda x, y, z=None: transform_coord(x, y), geom)
            return transformed
        except Exception:
            return None

    # Collect all tiles we need to generate
    print('  Computing tile coverage...')
    tiles = {}  # (z, x, y) → [(geom, props), ...]

    for geom, props in parsed:
        bounds = geom.bounds  # (minx, miny, maxx, maxy) = (min_lng, min_lat, max_lng, max_lat)
        for z in range(min_zoom, max_zoom + 1):
            # Find all tiles this geometry touches
            tx1, ty2 = lng_lat_to_tile(bounds[0], bounds[1], z)  # SW corner
            tx2, ty1 = lng_lat_to_tile(bounds[2], bounds[3], z)  # NE corner
            # Limit tile range to avoid huge loops at low zooms for large polygons
            max_tiles_per_axis = min(2 ** z, 64)
            if (tx2 - tx1 + 1) * (ty2 - ty1 + 1) > max_tiles_per_axis * max_tiles_per_axis:
                # Simplify: only add if geometry is large enough to matter at this zoom
                if z > 3:
                    continue
            for tx in range(tx1, tx2 + 1):
                for ty in range(ty1, ty2 + 1):
                    key = (z, tx, ty)
                    if key not in tiles:
                        tiles[key] = []
                    tiles[key].append((geom, props))

    print(f'  {len(tiles)} tiles to generate across z{min_zoom}-z{max_zoom}')

    # Generate MVT tiles and write PMTiles
    tile_data = {}  # (z, x, y) → bytes
    total = len(tiles)
    done = 0
    for (z, x, y), tile_features in tiles.items():
        bbox = tile_bounds(x, y, z)
        tile_box = box(*bbox)
        mvt_features = []
        for geom, props in tile_features:
            try:
                clipped = geom.intersection(tile_box)
                if clipped.is_empty:
                    continue
                # Simplify for lower zooms
                if z < 8:
                    tol = 360 / (2 ** z) / 4096 * 4
                    clipped = clipped.simplify(tol, preserve_topology=True)
                    if clipped.is_empty:
                        continue
                # Convert props — MVT needs simple types
                clean_props = {}
                for k, v in props.items():
                    if v is None:
                        continue
                    if isinstance(v, float):
                        if v != v:  # NaN check
                            continue
                        clean_props[k] = v
                    else:
                        clean_props[k] = str(v)
                mvt_features.append({
                    'geometry': mapping(clipped),
                    'properties': clean_props
                })
            except Exception:
                continue

        if not mvt_features:
            done += 1
            continue

        try:
            tile_bytes = mvt.encode([{
                'name': 'utility_territories',
                'features': mvt_features
            }], quantize_bounds=(bbox[0], bbox[1], bbox[2], bbox[3]))
            tile_data[(z, x, y)] = gzip.compress(tile_bytes)
        except Exception as e:
            pass
        done += 1
        if done % 500 == 0:
            print(f'  Generated {done}/{total} tiles...')

    print(f'  Generated {len(tile_data)} non-empty tiles')

    # Write PMTiles file
    print('  Writing PMTiles...')
    with open(OUT_PMTILES, 'wb') as f:
        writer = PMWriter(f)
        for (z, x, y), data in sorted(tile_data.items()):
            tileid = zxy_to_tileid(z, x, y)
            writer.write_tile(tileid, data)

        metadata = {
            'name': 'utility_territories',
            'description': 'US Electric Utility Service Territories (HIFLD + EIA 861 2024)',
            'format': 'pbf',
            'type': 'overlay',
            'minzoom': str(min_zoom),
            'maxzoom': str(max_zoom),
            'vector_layers': [{
                'id': 'utility_territories',
                'description': 'Electric utility service territory polygons',
                'fields': {
                    'utility_id': 'Number',
                    'name': 'String',
                    'state': 'String',
                    'type': 'String',
                    'holding_company': 'String',
                    'control_area': 'String',
                    'nerc_region': 'String',
                    'regulated': 'String',
                    'summer_peak_mw': 'Number',
                    'winter_peak_mw': 'Number',
                    'net_generation_mwh': 'Number',
                    'total_purchases_mwh': 'Number',
                    'total_customers': 'Number',
                    'website': 'String',
                    'telephone': 'String',
                    'data_year': 'String'
                }
            }]
        }
        writer.finalize(
            header={
                'tile_type': TileType.MVT,
                'tile_compression': Compression.GZIP,
                'min_zoom': min_zoom,
                'max_zoom': max_zoom,
                'min_lon_e7': -1800000000,
                'min_lat_e7': -900000000,
                'max_lon_e7': 1800000000,
                'max_lat_e7': 900000000,
            },
            metadata=metadata
        )

    size_mb = os.path.getsize(OUT_PMTILES) / 1024 / 1024
    print(f'  Wrote {OUT_PMTILES} ({size_mb:.1f} MB)')


if __name__ == '__main__':
    features = fetch_hifld()
    ops = parse_operational_data()
    customers = parse_sales_data()
    features = join_data(features, ops, customers)
    geojson = write_geojson(features)
    build_pmtiles(geojson)
    print('\nDone!')
