#!/usr/bin/env python3
"""
Process gas pipeline data files from EBB portals.

Usage:
    python refresh/process.py --unsub    # Process unsubscribed capacity files
    python refresh/process.py --ioc      # Process IOC contract files
    python refresh/process.py --tpit     # Process ERCOT TPIT planned transmission

Requirements: pip install openpyxl xlrd
"""

import argparse
import csv
import json
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

try:
    import xlrd
except ImportError:
    print("ERROR: pip install xlrd")
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parent.parent
REFRESH_DIR = REPO_ROOT / "refresh"
DATA_DIR = REPO_ROOT / "data"

CURRENT_YEAR = datetime.now().year
TODAY = datetime.now().strftime("%Y-%m-%d")

# ─── CONTENT-BASED PIPELINE IDENTIFICATION ───────────────────────────────────

# Patterns matched against file content (first ~50 rows) to identify pipeline
CONTENT_PATTERNS = [
    (r"Tennessee Gas Pipeline",                "Tennessee Gas Pipeline"),
    (r"Natural Gas Pipeline.*America|NGPL",    "NGPL"),
    (r"El Paso Natural Gas",                   "El Paso Natural Gas"),
    (r"Southern Natural Gas",                  "Southern Natural Gas"),
    (r"Colorado Interstate Gas",               "Colorado Interstate Gas"),
    (r"Midcontinent Express",                  "Midcontinent Express"),
    (r"Kinder Morgan",                         "Kinder Morgan"),
    (r"ANR Pipeline",                          "ANR Pipeline"),
    (r"Columbia Gas Transmission",             "Columbia Gas Transmission"),
    (r"Columbia Gulf Transmission",            "Columbia Gulf Transmission"),
    (r"Great Lakes Gas Transmission",          "Great Lakes Gas Transmission"),
    (r"TC Energy|TransCanada",                 "TC Energy"),
    (r"ROCKIES EXPRESS PIPELINE",              "Rockies Express"),
    (r"TRAILBLAZER PIPELINE",                  "Trailblazer Pipeline"),
    (r"Tallgrass",                             "Tallgrass"),
    (r"Vector Pipeline",                       "Vector Pipeline"),
    (r"Midwestern Gas Transmission",           "Midwestern Gas Transmission"),
    (r"DT Midstream|Millennium Pipeline",      "DT Midstream"),
    (r"Transcontinental Gas Pipe Line",        "Transcontinental Gas Pipe Line"),
    (r"Williams",                              "Williams"),
    (r"Northern Natural Gas",                  "Northern Natural Gas"),
    (r"Gulf South Pipeline",                   "Gulf South Pipeline"),
    (r"Texas Gas Transmission",                "Texas Gas Transmission"),
    (r"Boardwalk",                             "Boardwalk"),
    (r"NEXUS Gas Transmission",                "NEXUS Gas Transmission"),
    (r"Panhandle Eastern",                     "Panhandle Eastern"),
    (r"Trunkline",                             "Trunkline"),
    (r"Rover Pipeline",                        "Rover Pipeline"),
    (r"Enable",                                "Enable"),
    (r"Southern Star",                         "Southern Star"),
    (r"Enbridge",                              "Enbridge"),
]

# KM IC filename codes (e.g., IC0006232601.xls -> TGP)
KM_FILENAME_CODES = {
    "TGP": "Tennessee Gas Pipeline",
    "NGPL": "NGPL",
    "EPNG": "El Paso Natural Gas",
    "SNG": "Southern Natural Gas",
    "CIG": "Colorado Interstate Gas",
    "MEP": "Midcontinent Express",
}


def detect_format(filepath):
    """Detect file format from extension."""
    return filepath.suffix.lower().lstrip(".")


def read_xlsx(filepath, sheet_name=None):
    """Read xlsx file, return list of lists."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def read_xls(filepath, sheet_index=0):
    """Read old .xls file via xlrd, return list of lists."""
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(sheet_index)
    return [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]


def read_delimited(filepath, delimiter=None):
    """Read CSV/TSV/tab-delimited file."""
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        sample = f.read(4096)
        f.seek(0)
        if delimiter is None:
            delimiter = "\t" if "\t" in sample else ","
        return list(csv.reader(f, delimiter=delimiter))


def read_file(filepath):
    """Read any supported file format, return list of lists."""
    fmt = detect_format(filepath)
    if fmt == "xlsx":
        return read_xlsx(filepath)
    elif fmt == "xls":
        return read_xls(filepath)
    elif fmt == "csv":
        return read_delimited(filepath, delimiter=",")
    elif fmt in ("txt", "tab"):
        return read_delimited(filepath, delimiter="\t")
    else:
        return read_delimited(filepath)


def identify_pipeline_from_content(rows, filepath):
    """Identify pipeline by examining file content (first 50 rows).
    Falls back to KM filename codes, then returns None.
    """
    # Build a text blob from the first 50 rows
    text_blob = ""
    for row in rows[:50]:
        text_blob += " ".join(safe_str(c) for c in row) + "\n"

    for pattern, pipeline_name in CONTENT_PATTERNS:
        if re.search(pattern, text_blob, re.IGNORECASE):
            return pipeline_name

    # Try KM filename codes (e.g., IC0006232601.xls)
    stem = filepath.stem.upper()
    for code, pipeline_name in KM_FILENAME_CODES.items():
        if code in stem:
            return pipeline_name

    return None


def find_header_row(rows, keywords):
    """Find the row index containing header keywords."""
    for i, row in enumerate(rows):
        row_str = " ".join(str(c).lower() for c in row if c)
        matches = sum(1 for kw in keywords if kw.lower() in row_str)
        if matches >= 2:
            return i
    return 0


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    return int(safe_float(val, default))


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def parse_date_str(val):
    """Normalize a date value to MM/DD/YYYY string."""
    s = safe_str(val)
    if not s:
        return ""
    # Already in MM/DD/YYYY
    if re.match(r"\d{2}/\d{2}/\d{4}", s):
        return s
    # YYYY-MM-DD
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(2)}/{m.group(3)}/{m.group(1)}"
    # Try datetime objects from Excel
    if hasattr(val, "strftime"):
        return val.strftime("%m/%d/%Y")
    # xlrd float date
    if isinstance(val, float) and val > 30000:
        try:
            dt = xlrd.xldate_as_datetime(val, 0)
            return dt.strftime("%m/%d/%Y")
        except Exception:
            pass
    return s


def contract_status(end_date_str):
    """Determine status: expired, expiring_soon, or active."""
    if not end_date_str:
        return "active"
    # Extract year from end date
    m = re.search(r"(\d{4})", end_date_str)
    if not m:
        return "active"
    end_year = int(m.group(1))
    if end_year < CURRENT_YEAR:
        return "expired"
    elif end_year <= CURRENT_YEAR + 1:
        return "expiring_soon"
    return "active"


# ─── UNSUB PROCESSING ───────────────────────────────────────────────────────

def parse_unsub_naesb(rows):
    """NAESB format (Tallgrass, Vector-like): find LOC_ID and UNSUB columns."""
    header_idx = find_header_row(rows, ["LOC", "UNSUB"])
    if header_idx >= len(rows):
        header_idx = find_header_row(rows, ["Location", "Unsub"])
    headers = [safe_str(h).upper() for h in rows[header_idx]]

    loc_col = None
    unsub_col = None
    for i, h in enumerate(headers):
        if "LOC_ID" in h or h == "LOCATION" or (h == "LOC" and loc_col is None):
            loc_col = i
        if "UNSUB" in h:
            unsub_col = i

    if loc_col is None or unsub_col is None:
        return {}

    points = {}
    for row in rows[header_idx + 1:]:
        if len(row) <= max(loc_col, unsub_col):
            continue
        loc_id = safe_str(row[loc_col])
        qty = safe_int(row[unsub_col])
        if loc_id and qty > 0:
            points[loc_id] = qty
    return points


def parse_unsub_vector(rows):
    """Vector format: TSP Name, TSP, Location, Location_Name, Loc_Purp, Loc/QTI, Unsub_Cap."""
    header_idx = find_header_row(rows, ["TSP", "Unsub"])
    headers = [safe_str(h) for h in rows[header_idx]]

    loc_col = None
    unsub_col = None
    for i, h in enumerate(headers):
        if "Location" == h or "Loc_ID" in h:
            if loc_col is None:
                loc_col = i
        if "Unsub" in h:
            unsub_col = i

    if loc_col is None:
        loc_col = 2
    if unsub_col is None:
        unsub_col = 6

    points = {}
    for row in rows[header_idx + 1:]:
        if len(row) <= max(loc_col, unsub_col):
            continue
        loc_id = safe_str(row[loc_col])
        qty = safe_int(row[unsub_col])
        if loc_id and qty > 0:
            points[loc_id] = qty
    return points


def parse_unsub_et(rows):
    """ET Messenger: Loc, Loc Name, Loc Purp Desc, Loc/QTI, Unsubscribed Capacity."""
    header_idx = find_header_row(rows, ["Loc", "Unsubscribed"])
    points = {}
    for row in rows[header_idx + 1:]:
        if len(row) < 5:
            continue
        loc_id = safe_str(row[0])
        qty = safe_int(row[4])
        if loc_id and qty > 0:
            points[loc_id] = qty
    return points


def parse_unsub_generic(rows):
    """Generic unsub parser — detect loc and capacity columns by header names."""
    header_idx = find_header_row(rows, ["loc", "capacity"])
    if header_idx >= len(rows):
        return {}
    headers = [safe_str(h).lower() for h in rows[header_idx]]

    loc_col = None
    unsub_col = None
    for i, h in enumerate(headers):
        if ("loc" in h and "id" in h) or h in ("loc", "location", "point"):
            if loc_col is None:
                loc_col = i
        if "unsub" in h or "avail" in h or "capacity" in h:
            if unsub_col is None:
                unsub_col = i

    if loc_col is None:
        loc_col = 0
    if unsub_col is None:
        unsub_col = len(headers) - 1

    points = {}
    for row in rows[header_idx + 1:]:
        if len(row) <= max(loc_col, unsub_col):
            continue
        loc_id = safe_str(row[loc_col])
        qty = safe_int(row[unsub_col])
        if loc_id and qty > 0:
            points[loc_id] = qty
    return points


def detect_unsub_format(rows):
    """Detect which unsub parser to use based on headers."""
    first_rows_text = " ".join(
        safe_str(c) for row in rows[:10] for c in row
    ).upper()

    if "TSP_NM" in first_rows_text or "TSP_NO" in first_rows_text:
        return "naesb"
    if "TSP NAME" in first_rows_text or "UNSUB_CAP" in first_rows_text:
        return "vector"
    if "LOC PURP DESC" in first_rows_text or "UNSUBSCRIBED CAPACITY" in first_rows_text:
        return "et"
    return "generic"


def process_unsub():
    """Process all files in refresh/unsub/ and output data/unsub_capacity.json."""
    unsub_dir = REFRESH_DIR / "unsub"
    if not unsub_dir.exists():
        print(f"ERROR: {unsub_dir} does not exist")
        return None

    files = [f for f in unsub_dir.iterdir() if f.is_file() and not f.name.startswith((".", "~"))]
    if not files:
        print("No files found in refresh/unsub/")
        return None

    # Load existing data for merge
    output_path = DATA_DIR / "unsub_capacity.json"
    existing = {}
    if output_path.exists():
        with open(output_path, "r") as f:
            data = json.load(f)
            existing = data.get("pipelines", {})

    prev_counts = {p: len(pts) for p, pts in existing.items()}
    updated_pipelines = {}

    for filepath in sorted(files):
        fmt = detect_format(filepath)
        print(f"  Reading {filepath.name} ({fmt})")

        try:
            rows = read_file(filepath)
        except Exception as e:
            print(f"    ERROR reading file: {e}")
            continue

        if not rows:
            print("    WARNING: empty file")
            continue

        pipeline = identify_pipeline_from_content(rows, filepath)
        if pipeline is None:
            print(f"    WARNING: Could not identify pipeline from content.")
            print(f"    Skipping {filepath.name} — please rename or identify manually.")
            continue

        print(f"    Identified: {pipeline}")

        # Parse based on detected format
        fmt_type = detect_unsub_format(rows)
        if fmt_type == "naesb":
            points = parse_unsub_naesb(rows)
        elif fmt_type == "vector":
            points = parse_unsub_vector(rows)
        elif fmt_type == "et":
            points = parse_unsub_et(rows)
        else:
            points = parse_unsub_generic(rows)

        if points:
            updated_pipelines[pipeline] = points
            print(f"    Parsed {len(points)} points")
        else:
            print(f"    WARNING: no points parsed")

    # Merge: update only pipelines with new files, preserve the rest
    merged = dict(existing)
    for pipeline, points in updated_pipelines.items():
        merged[pipeline] = points

    output = {
        "generated": TODAY,
        "pipelines": merged,
    }

    with open(output_path, "w") as f:
        json.dump(output, f, indent=2)

    # Summary
    total_points = sum(len(pts) for pts in merged.values())
    print(f"\n{'='*60}")
    print(f"UNSUB SUMMARY — {total_points} total points across {len(merged)} pipelines")
    print(f"{'='*60}")
    for pipe in sorted(updated_pipelines.keys()):
        count = len(updated_pipelines[pipe])
        prev = prev_counts.get(pipe, 0)
        delta = count - prev
        delta_str = f" ({'+' if delta > 0 else ''}{delta})" if prev > 0 else " (new)"
        print(f"  {pipe:40s}  {count:5d} points{delta_str}")
    preserved = set(merged.keys()) - set(updated_pipelines.keys())
    if preserved:
        print(f"  --- preserved from previous run ---")
        for pipe in sorted(preserved):
            print(f"  {pipe:40s}  {len(merged[pipe]):5d} points (unchanged)")
    print(f"\nWritten to {output_path}")
    return "unsub"


# ─── IOC PROCESSING ─────────────────────────────────────────────────────────

def parse_ioc_tc_energy(rows, pipeline):
    """TC Energy H/D/P Excel format.
    Cols: 1=type, 5=shipper, 14=rate_sched, 17=contract, 18=begin, 19=end, 25=trans_mdq, 28=stor_mdq.
    """
    contracts = []
    for row in rows[1:]:
        if len(row) < 29:
            continue
        rec_type = safe_str(row[1])
        if rec_type not in ("H", "D", "A", "P"):
            continue

        end_date = parse_date_str(row[19])
        mdq = safe_int(row[25]) + safe_int(row[28])

        contracts.append({
            "shipper": safe_str(row[5]),
            "contract": safe_str(row[17]),
            "rate_schedule": safe_str(row[14]),
            "begin": parse_date_str(row[18]),
            "end": end_date,
            "mdq": mdq,
            "status": contract_status(end_date),
            "points": [],  # TC Energy format doesn't include point details in H record
        })
    return contracts


def parse_ioc_hdp_generic(rows):
    """Generic H/D/A/P IOC format used by KM, Williams, Tallgrass, Vector.
    Returns list of contract dicts. Groups D/P point records under their parent H record.
    """
    header_idx = find_header_row(rows, ["shipper", "contract", "rate"])
    if header_idx >= len(rows):
        header_idx = 0

    headers = [safe_str(h).lower() for h in rows[header_idx]]

    # Find key columns
    col = {}
    for i, h in enumerate(headers):
        if "type" in h or "rec" in h:
            col.setdefault("type", i)
        elif "shipper" in h:
            col.setdefault("shipper", i)
        elif "rate" in h and "sched" in h:
            col.setdefault("rate_schedule", i)
        elif "contract" in h:
            col.setdefault("contract", i)
        elif "begin" in h or "start" in h or "eff" in h:
            col.setdefault("begin", i)
        elif "end" in h or "expir" in h:
            col.setdefault("end", i)
        elif "mdq" in h or "max" in h and "quant" in h:
            col.setdefault("mdq", i)
        elif "point" in h and ("name" in h or "nm" in h):
            col.setdefault("point_name", i)
        elif "point" in h and ("id" in h or "no" in h or "num" in h):
            col.setdefault("point_id", i)
        elif "zone" in h:
            col.setdefault("zone", i)

    def get(row, key, default=""):
        idx = col.get(key)
        if idx is not None and idx < len(row):
            return row[idx]
        return default

    contracts = []
    current_h = None

    for row in rows[header_idx + 1:]:
        if not row or len(row) < 3:
            continue

        rec_type = safe_str(get(row, "type"))
        if not rec_type:
            rec_type = safe_str(row[0]) if row else ""

        if rec_type == "H":
            # Save previous H record
            if current_h is not None:
                contracts.append(current_h)

            end_date = parse_date_str(get(row, "end"))
            current_h = {
                "shipper": safe_str(get(row, "shipper")),
                "contract": safe_str(get(row, "contract")),
                "rate_schedule": safe_str(get(row, "rate_schedule")),
                "begin": parse_date_str(get(row, "begin")),
                "end": end_date,
                "mdq": safe_int(get(row, "mdq")),
                "status": contract_status(end_date),
                "points": [],
            }

        elif rec_type in ("D", "A", "P") and current_h is not None:
            point = {
                "type": rec_type,
                "name": safe_str(get(row, "point_name")),
                "id": safe_str(get(row, "point_id")),
                "zone": safe_str(get(row, "zone")),
                "mdq": safe_int(get(row, "mdq")),
            }
            # Use contract-level name columns as fallback for point name
            if not point["name"] and "shipper" in col and col["shipper"] < len(row):
                # Some formats reuse shipper column for point name on D rows
                candidate = safe_str(row[col["shipper"]])
                if candidate and candidate != current_h["shipper"]:
                    point["name"] = candidate
            current_h["points"].append(point)

    # Don't forget the last H record
    if current_h is not None:
        contracts.append(current_h)

    return contracts


def build_pipeline_ioc(pipeline, contracts):
    """Build pipeline-level IOC structure with summary stats."""
    total = len(contracts)
    expiring = [c for c in contracts if c.get("status") == "expiring_soon"]
    expiring_count = len(expiring)
    expiring_mdq = sum(c.get("mdq", 0) for c in expiring)

    return {
        "total_contracts": total,
        "expiring_count": expiring_count,
        "expiring_mdq": expiring_mdq,
        "contracts": contracts,
    }


def process_ioc():
    """Process all files in refresh/ioc/ and output data/ioc_contracts.json."""
    ioc_dir = REFRESH_DIR / "ioc"
    if not ioc_dir.exists():
        print(f"ERROR: {ioc_dir} does not exist")
        return None

    files = [f for f in ioc_dir.iterdir() if f.is_file() and not f.name.startswith((".", "~"))]
    if not files:
        print("No files found in refresh/ioc/")
        return None

    # Load existing data for merge
    output_path = DATA_DIR / "ioc_contracts.json"
    existing = {}
    if output_path.exists():
        with open(output_path, "r") as f:
            data = json.load(f)
            existing = data.get("pipelines", {})

    prev_counts = {p: v.get("total_contracts", 0) for p, v in existing.items()}
    updated_pipelines = {}

    for filepath in sorted(files):
        fmt = detect_format(filepath)
        print(f"  Reading {filepath.name} ({fmt})")

        try:
            rows = read_file(filepath)
        except Exception as e:
            print(f"    ERROR reading file: {e}")
            continue

        if not rows:
            print("    WARNING: empty file")
            continue

        pipeline = identify_pipeline_from_content(rows, filepath)
        if pipeline is None:
            print(f"    WARNING: Could not identify pipeline from content.")
            print(f"    Skipping {filepath.name} — please rename or identify manually.")
            continue

        print(f"    Identified: {pipeline}")

        # Detect if it's TC Energy column layout
        first_text = " ".join(safe_str(c) for row in rows[:5] for c in row).lower()
        is_tc = bool(re.search(r"anr|columbia|great lakes|tc energy|transcanada", first_text))

        if is_tc and fmt in ("xlsx", "xls"):
            contracts = parse_ioc_tc_energy(rows, pipeline)
        else:
            contracts = parse_ioc_hdp_generic(rows)

        if contracts:
            updated_pipelines[pipeline] = build_pipeline_ioc(pipeline, contracts)
            print(f"    Parsed {len(contracts)} contracts")
        else:
            print(f"    WARNING: no contracts parsed")

    # Merge: update only pipelines with new files, preserve the rest
    merged = dict(existing)
    for pipeline, ioc_data in updated_pipelines.items():
        merged[pipeline] = ioc_data

    # Build global summary
    total_contracts = sum(v.get("total_contracts", 0) for v in merged.values())
    total_pipelines = len(merged)
    expiring_all = sum(v.get("expiring_count", 0) for v in merged.values())
    expiring_mdq_all = sum(v.get("expiring_mdq", 0) for v in merged.values())

    output = {
        "generated": TODAY,
        "pipelines": merged,
        "summary": {
            "total_contracts": total_contracts,
            "total_pipelines": total_pipelines,
            f"expiring_{CURRENT_YEAR}_{CURRENT_YEAR+1}": expiring_all,
            "expiring_mdq_dth": expiring_mdq_all,
        },
    }

    with open(output_path, "w") as f:
        json.dump(output, f, indent=2)

    # Summary
    print(f"\n{'='*60}")
    print(f"IOC SUMMARY — {total_contracts} contracts across {total_pipelines} pipelines")
    print(f"{'='*60}")
    for pipe in sorted(updated_pipelines.keys()):
        count = updated_pipelines[pipe]["total_contracts"]
        exp = updated_pipelines[pipe]["expiring_count"]
        prev = prev_counts.get(pipe, 0)
        delta = count - prev
        delta_str = f" ({'+' if delta > 0 else ''}{delta})" if prev > 0 else " (new)"
        print(f"  {pipe:40s}  {count:5d} contracts, {exp:4d} expiring{delta_str}")
    preserved = set(merged.keys()) - set(updated_pipelines.keys())
    if preserved:
        print(f"  --- preserved from previous run ---")
        for pipe in sorted(preserved):
            c = merged[pipe].get("total_contracts", 0)
            print(f"  {pipe:40s}  {c:5d} contracts (unchanged)")
    print(f"\n  Expiring {CURRENT_YEAR}-{CURRENT_YEAR+1}: {expiring_all} contracts, {expiring_mdq_all:,} Dth/d MDQ")
    print(f"\nWritten to {output_path}")
    return "ioc"


# ─── TPIT PROCESSING ────────────────────────────────────────────────────────

def process_tpit():
    """Process ERCOT TPIT xlsx and merge with planned_transmission.json."""
    tpit_dir = REFRESH_DIR / "tpit"
    if not tpit_dir.exists():
        print(f"ERROR: {tpit_dir} does not exist")
        return None

    xlsx_files = [f for f in tpit_dir.iterdir() if f.suffix.lower() == ".xlsx" and not f.name.startswith("~")]
    if not xlsx_files:
        print("No .xlsx files found in refresh/tpit/")
        return None

    tpit_file = xlsx_files[0]
    print(f"  Reading {tpit_file.name}")

    # Load geocoding reference
    geo_path = DATA_DIR / "capacity_screening.json"
    geo_lookup = {}
    if geo_path.exists():
        with open(geo_path, "r") as f:
            geo_data = json.load(f)
            for item in geo_data:
                name = safe_str(item.get("name", "")).lower()
                if name and "latitude" in item and "longitude" in item:
                    geo_lookup[name] = {
                        "latitude": item["latitude"],
                        "longitude": item["longitude"],
                    }
        print(f"    Loaded {len(geo_lookup)} geocoded substations")
    else:
        print("    WARNING: capacity_screening.json not found, skipping geocoding")

    # Load existing planned_transmission.json
    output_path = DATA_DIR / "planned_transmission.json"
    existing = []
    existing_keys = set()
    if output_path.exists():
        with open(output_path, "r") as f:
            existing = json.load(f)
            for proj in existing:
                key = (
                    safe_str(proj.get("project_name", "")).lower(),
                    safe_str(proj.get("from_bus", "")).lower(),
                    safe_str(proj.get("to_bus", "")).lower(),
                )
                existing_keys.add(key)
    prev_count = len(existing)

    # Parse TPIT sheets
    wb = openpyxl.load_workbook(tpit_file, read_only=True, data_only=True)
    new_projects = []

    for sheet_name in ["FutureTPIT", "PlannedTPIT"]:
        if sheet_name not in wb.sheetnames:
            print(f"    WARNING: sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [safe_str(h).lower() for h in rows[0]]

        col_map = {}
        for i, h in enumerate(headers):
            if "project" in h and "name" in h:
                col_map["project_name"] = i
            elif "from" in h and ("bus" in h or "station" in h or "sub" in h):
                col_map["from_bus"] = i
            elif "to" in h and ("bus" in h or "station" in h or "sub" in h):
                col_map["to_bus"] = i
            elif h in ("voltage", "kv", "voltage (kv)"):
                col_map["voltage"] = i
            elif "status" in h:
                col_map["status"] = i
            elif "type" in h:
                col_map["type"] = i
            elif "in" in h and "service" in h:
                col_map["in_service"] = i
            elif "county" in h or "counties" in h:
                col_map["county"] = i

        def get_col(row, key):
            idx = col_map.get(key)
            if idx is not None and idx < len(row):
                return safe_str(row[idx])
            return ""

        sheet_count = 0
        for row in rows[1:]:
            if not row or all(c is None for c in row):
                continue

            project_name = get_col(row, "project_name")
            from_bus = get_col(row, "from_bus")
            to_bus = get_col(row, "to_bus")

            if not project_name:
                continue

            key = (project_name.lower(), from_bus.lower(), to_bus.lower())
            if key in existing_keys:
                continue

            project = {
                "project_name": project_name,
                "from_bus": from_bus,
                "to_bus": to_bus,
                "source": sheet_name,
            }

            for field in ("voltage", "status", "type", "county"):
                val = get_col(row, field)
                if val:
                    project[field] = val
            in_svc = get_col(row, "in_service")
            if in_svc:
                project["in_service_date"] = in_svc

            # Geocode
            from_key = from_bus.lower()
            to_key = to_bus.lower()
            if from_key in geo_lookup:
                project["from_lat"] = geo_lookup[from_key]["latitude"]
                project["from_lon"] = geo_lookup[from_key]["longitude"]
            if to_key in geo_lookup:
                project["to_lat"] = geo_lookup[to_key]["latitude"]
                project["to_lon"] = geo_lookup[to_key]["longitude"]

            new_projects.append(project)
            existing_keys.add(key)
            sheet_count += 1

        print(f"    {sheet_name}: {sheet_count} new projects")

    wb.close()

    merged = existing + new_projects

    with open(output_path, "w") as f:
        json.dump(merged, f, indent=2)

    geocoded = sum(1 for p in new_projects if "from_lat" in p or "to_lat" in p)

    print(f"\n{'='*60}")
    print(f"TPIT SUMMARY")
    print(f"{'='*60}")
    print(f"  Previous projects:  {prev_count}")
    print(f"  New projects added: {len(new_projects)}")
    print(f"  Geocoded (partial): {geocoded}")
    print(f"  Total projects:     {len(merged)}")
    print(f"\nWritten to {output_path}")
    return "tpit"


# ─── GIT COMMIT ──────────────────────────────────────────────────────────────

def git_commit_and_push(refresh_type):
    """Stage data/, commit, and push."""
    today = datetime.now().strftime("%Y-%m-%d")
    msg = f"refresh: {refresh_type} {today}"
    print(f"\nCommitting: {msg}")
    try:
        subprocess.run(["git", "add", "data/"], cwd=REPO_ROOT, check=True)
        subprocess.run(["git", "commit", "-m", msg], cwd=REPO_ROOT, check=True)
        subprocess.run(["git", "push"], cwd=REPO_ROOT, check=True)
        print("Pushed to remote.")
    except subprocess.CalledProcessError as e:
        print(f"Git error: {e}")
        print("You may need to commit and push manually.")


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Process gas pipeline data files")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--unsub", action="store_true", help="Process unsubscribed capacity")
    group.add_argument("--ioc", action="store_true", help="Process IOC contracts")
    group.add_argument("--tpit", action="store_true", help="Process ERCOT TPIT planned transmission")
    parser.add_argument("--no-push", action="store_true", help="Skip git commit and push")
    args = parser.parse_args()

    print(f"\n{'='*60}")
    if args.unsub:
        print("Processing UNSUBSCRIBED CAPACITY")
        print(f"{'='*60}")
        result = process_unsub()
    elif args.ioc:
        print("Processing IOC CONTRACTS")
        print(f"{'='*60}")
        result = process_ioc()
    elif args.tpit:
        print("Processing ERCOT TPIT")
        print(f"{'='*60}")
        result = process_tpit()

    if result and not args.no_push:
        git_commit_and_push(result)


if __name__ == "__main__":
    main()
